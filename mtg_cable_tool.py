import os
import re
import math
import ctypes
import sys
from datetime import date
import tkinter as tk
import tkinter.font as tkfont
from tkinter import filedialog, messagebox, ttk, simpledialog

from openpyxl import Workbook, load_workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


APP_NAME = "MTG Cable Tool"
APP_VERSION = "1.3.3"
EXTRACT_BLANK_ROW_STOP = 75

# Union of worksheet source fields used anywhere in the application.
# These are pre-cached when a sheet is first scanned so Single File,
# Multiple Files / Sheets, and Cable Lengths can share the same raw data.
SHARED_SOURCE_FIELDS = (
    "a_position",
    "z_position",
    "cable_type",
    "path",
)

OUTPUT_HEADERS = [
    "A Position",
    "Project Number",
    "Path",
    "Cable Number",
    "Identifier",
    "Other",
    "Cable Type",
    "Z Position",
]


def normalize_header(value):
    if value is None:
        return ""
    return re.sub(r"[\s_\-]+", "", str(value).strip().lower())


def clean_value(value):
    if value is None:
        return ""
    return str(value).strip()




def simplify_cable_type(value):
    original = clean_value(value)
    text = original.upper()

    # Keep copper descriptions exactly as they appear in the source cell.
    if "COPPER" in text:
        return original

    # For non-copper cables, keep the fiber count and connector family only.
    fiber_match = re.search(r"\b(\d+F)\b", text)
    if fiber_match:
        fiber_count = fiber_match.group(1)

        if "LC" in text:
            return f"{fiber_count} LC"

        return fiber_count

    if "LC" in text:
        return "LC"

    return original

def header_tokens(value):
    if value is None:
        return []

    text = str(value).strip().lower()
    text = re.sub(r"[_\-]+", " ", text)
    text = re.sub(r"[^a-z0-9 ]+", " ", text)
    return [token for token in text.split() if token]


def normalized_header(value):
    return "".join(header_tokens(value))


def header_matches(value, key):
    normalized = normalized_header(value)

    if not normalized:
        return False

    aliases = {
        "a_position": {
            "aposition",
            "apos",
            "positiona",
        },
        "z_position": {
            "zposition",
            "zpos",
            "positionz",
        },
        "cable_type": {
            "cabletype",
            "fibertype",
        },
        "path": {
            "path",
            "cablepath",
        },
    }

    return normalized in aliases.get(key, set())


def find_header_columns(ws):
    keys = ("a_position", "z_position", "cable_type", "path")

    # Score each candidate row. Required headers should normally be together.
    best = None

    for row_num in range(1, min(ws.max_row, 75) + 1):
        found = {}

        for col_num in range(1, ws.max_column + 1):
            value = ws.cell(row=row_num, column=col_num).value

            for key in keys:
                if key not in found and header_matches(value, key):
                    found[key] = col_num

        required_count = sum(
            key in found
            for key in ("a_position", "z_position", "cable_type")
        )

        if required_count == 3:
            # Prefer a row that also contains Path.
            score = 100 + (1 if "path" in found else 0)
            if best is None or score > best[0]:
                best = (score, row_num, found)

    if best is None:
        return None, None

    return best[1], best[2]


def describe_detected_columns(ws, columns):
    def describe(key, label):
        col_num = columns.get(key)

        if not col_num:
            return f"{label}: not found"

        letter = get_column_letter(col_num)
        header_value = clean_value(ws.cell(row=1, column=col_num).value)
        return f"{label}: {letter}"

    parts = [
        describe("a_position", "A Position"),
        describe("cable_type", "Cable Type"),
        describe("z_position", "Z Position"),
        describe("path", "Path"),
    ]

    return " | ".join(parts)


def extract_labels(
    file_path,
    sheet_name=None,
    header_row_override=None,
    columns_override=None,
    progress_callback=None,
):
    workbook = load_workbook(
        file_path,
        data_only=True,
        read_only=True,
    )

    try:
        if sheet_name:
            if sheet_name not in workbook.sheetnames:
                raise RuntimeError(
                    f'Worksheet "{sheet_name}" was not found.'
                )
            worksheets = [workbook[sheet_name]]
        else:
            worksheets = list(workbook.worksheets)

        labels = []

        for ws in worksheets:
            if progress_callback:
                progress_callback()

            if columns_override:
                columns = dict(columns_override)
                header_row = header_row_override or 1
            else:
                header_row, columns = find_header_columns(ws)

            if not columns:
                continue

            required = (
                "a_position",
                "z_position",
                "cable_type",
            )

            if not all(columns.get(key) for key in required):
                continue

            a_col = columns["a_position"]
            z_col = columns["z_position"]
            cable_col = columns["cable_type"]
            path_col = columns.get("path")

            # Only inspect columns that are actually needed.
            needed_cols = {
                a_col,
                z_col,
                cable_col,
            }

            if path_col:
                needed_cols.add(path_col)

            min_col = min(needed_cols)
            max_col = max(needed_cols)

            consecutive_blank_rows = 0
            data_started = False

            # Stop after a sustained blank run once cable data has begun.
            BLANK_ROW_STOP = EXTRACT_BLANK_ROW_STOP

            for row_num in range(
                header_row + 1,
                ws.max_row + 1,
            ):
                if progress_callback and row_num % 50 == 0:
                    progress_callback()

                # Read only the mapped cells rather than the full worksheet row.
                a_value = ws.cell(
                    row=row_num,
                    column=a_col,
                ).value

                z_value = ws.cell(
                    row=row_num,
                    column=z_col,
                ).value

                cable_value = ws.cell(
                    row=row_num,
                    column=cable_col,
                ).value

                path_value = (
                    ws.cell(
                        row=row_num,
                        column=path_col,
                    ).value
                    if path_col
                    else ""
                )

                a_position = clean_value(a_value)
                z_position = clean_value(z_value)
                cable_type = clean_value(cable_value)
                path = clean_value(path_value)

                # "Blank" is based only on the columns relevant to this tool.
                row_is_blank = not any(
                    (
                        a_position,
                        z_position,
                        cable_type,
                        path,
                    )
                )

                if row_is_blank:
                    if data_started:
                        consecutive_blank_rows += 1

                        if consecutive_blank_rows >= BLANK_ROW_STOP:
                            break

                    continue

                data_started = True
                consecutive_blank_rows = 0

                # Cable Type marks the start of an actual cable row.
                # Blank cable type rows are continuation rows and are skipped.
                if not cable_type:
                    continue

                labels.append(
                    {
                        "a_position": a_position,
                        "z_position": z_position,
                        "cable_type": simplify_cable_type(
                            cable_type
                        ),
                        "path": path,
                        "source_sheet": ws.title,
                        "source_row": row_num,
                    }
                )

        return labels

    finally:
        workbook.close()


def write_output(
    output_path,
    labels,
    project_number,
    starting_cable_number,
    identifier,
    other,
):
    wb = Workbook()
    ws = wb.active
    ws.title = "Labels"

    ws.append(OUTPUT_HEADERS)

    for index, label in enumerate(labels):
        cable_number = starting_cable_number + index

        ws.append(
            [
                label["a_position"],
                project_number,
                label["path"],
                cable_number,
                identifier,
                other,
                label["cable_type"],
                label["z_position"],
            ]
        )

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    widths = [20, 18, 14, 15, 12, 15, 22, 24, 20]

    for col_num, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_num)].width = width

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="center")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(output_path)


def write_printable_output(
    output_path,
    labels,
    project_number,
    starting_cable_number,
    identifier,
    other,
):
    wb = Workbook()
    ws = wb.active
    ws.title = "Printable Labels"

    # Each cable label is stored in ONE merged cell.
    # Layout inside the cell:
    #
    # Path | Project Number | Identifier | Other
    # A Position | Cable Number | Cable Type | Z Position

    for col in ("A", "B", "C", "D"):
        ws.column_dimensions[col].width = 24

    thin = Side(style="thin", color="000000")
    label_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for index, label in enumerate(labels):
        row_num = index + 1
        cable_number = starting_cable_number + index

        top_line = (
            f"{label['path']:<20}"
            f"{project_number:<20}"
            f"{identifier:<20}"
            f"{other}"
        )

        bottom_line = (
            f"{label['a_position']:<20}"
            f"{str(cable_number):<20}"
            f"{label['cable_type']:<20}"
            f"{label['z_position']}"
        )

        label_text = f"{top_line}\n{bottom_line}"

        ws.merge_cells(
            start_row=row_num,
            start_column=1,
            end_row=row_num,
            end_column=4,
        )

        cell = ws.cell(row=row_num, column=1)
        cell.value = label_text
        cell.font = Font(name="Courier New", size=12)
        cell.alignment = Alignment(
            horizontal="left",
            vertical="center",
            wrap_text=True,
        )

        # Border around the full merged label area.
        for col_num in range(1, 5):
            ws.cell(row=row_num, column=col_num).border = label_border

        ws.row_dimensions[row_num].height = 38

    ws.sheet_view.showGridLines = False

    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_options.horizontalCentered = True

    margin = 0.25
    ws.page_margins.left = margin
    ws.page_margins.right = margin
    ws.page_margins.top = margin
    ws.page_margins.bottom = margin
    ws.page_margins.header = 0
    ws.page_margins.footer = 0

    wb.save(output_path)


def write_easymark_output(
    output_path,
    labels,
    project_number,
    starting_cable_number,
    identifier,
    other,
):
    wb = Workbook()
    ws = wb.active
    ws.title = "EasyMark"

    ws.column_dimensions["A"].width = 70

    bold_font = InlineFont(b=True)
    normal_font = InlineFont()

    for index, label in enumerate(labels):
        row_num = index + 1
        cable_number = starting_cable_number + index

        path_value = label["path"]
        first_line = (
            f"{cable_number} - {path_value}"
            if path_value
            else str(cable_number)
        )

        project_identifier = project_number
        if project_number and identifier:
            project_identifier = f"{project_number} - {identifier}"
        elif identifier:
            project_identifier = identifier

        # One Excel cell = one Easy-Mark label.
        # First three lines are bold.
        rich_text = CellRichText(
            TextBlock(bold_font, first_line),
            "\n",
            TextBlock(bold_font, label["a_position"]),
            "\n",
            TextBlock(bold_font, label["z_position"]),
            "\n",
            TextBlock(normal_font, label["cable_type"]),
            "\n",
            TextBlock(normal_font, project_identifier),
            "\n",
            TextBlock(normal_font, other),
        )

        cell = ws.cell(row=row_num, column=1)
        cell.value = rich_text
        cell.alignment = Alignment(
            horizontal="left",
            vertical="center",
            wrap_text=True,
        )

        ws.row_dimensions[row_num].height = 84

    ws.sheet_view.showGridLines = False
    wb.save(output_path)



def make_section_header(item):
    return (
        f"{item['sheet']} | "
        f"{item.get('project_number', '')} | "
        f"{item.get('identifier', '')} | "
        f"{item.get('other', '')} | "
        f"{item.get('file_name', '')}"
    )


def write_combined_output(output_path, sections):
    wb = Workbook()
    ws = wb.active
    ws.title = "Labels"

    row_num = 1

    for section_index, section in enumerate(sections):
        item = section["item"]
        labels = section["labels"]

        if section_index > 0:
            row_num += 1

        # Section header
        ws.merge_cells(
            start_row=row_num,
            start_column=1,
            end_row=row_num,
            end_column=len(OUTPUT_HEADERS),
        )
        header_cell = ws.cell(row=row_num, column=1)
        header_cell.value = make_section_header(item)
        header_cell.font = Font(bold=True)
        header_cell.alignment = Alignment(vertical="center")
        row_num += 1

        # Column headers
        for col_num, header in enumerate(OUTPUT_HEADERS, start=1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        row_num += 1

        start_number = int(item.get("cable_start", 1))

        for index, label in enumerate(labels):
            ws.cell(row=row_num, column=1).value = label["a_position"]
            ws.cell(row=row_num, column=2).value = item.get("project_number", "")
            ws.cell(row=row_num, column=3).value = start_number + index
            ws.cell(row=row_num, column=4).value = item.get("identifier", "")
            ws.cell(row=row_num, column=5).value = item.get("other", "")
            ws.cell(row=row_num, column=6).value = label["cable_type"]
            ws.cell(row=row_num, column=7).value = label["z_position"]
            ws.cell(row=row_num, column=8).value = label["path"]
            row_num += 1

    widths = [20, 18, 15, 18, 22, 38, 20, 14]
    for col_num, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_num)].width = width

    wb.save(output_path)


def write_combined_printable_output(output_path, sections):
    wb = Workbook()
    ws = wb.active
    ws.title = "Printable Labels"

    for col in ("A", "B", "C", "D"):
        ws.column_dimensions[col].width = 24

    thin = Side(style="thin", color="000000")
    label_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    row_num = 1

    for section_index, section in enumerate(sections):
        item = section["item"]
        labels = section["labels"]

        if section_index > 0:
            row_num += 1

        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=4)
        section_cell = ws.cell(row=row_num, column=1)
        section_cell.value = make_section_header(item)
        section_cell.font = Font(bold=True, size=12)
        section_cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.row_dimensions[row_num].height = 24
        row_num += 1

        start_number = int(item.get("cable_start", 1))

        for index, label in enumerate(labels):
            cable_number = start_number + index

            top_line = (
                f"{label['path']:<20}"
                f"{item.get('project_number', ''):<20}"
                f"{item.get('identifier', ''):<20}"
                f"{item.get('other', '')}"
            )
            bottom_line = (
                f"{label['a_position']:<20}"
                f"{str(cable_number):<20}"
                f"{label['cable_type']:<20}"
                f"{label['z_position']}"
            )

            ws.merge_cells(
                start_row=row_num,
                start_column=1,
                end_row=row_num,
                end_column=4,
            )

            cell = ws.cell(row=row_num, column=1)
            cell.value = f"{top_line}\n{bottom_line}"
            cell.font = Font(name="Courier New", size=12)
            cell.alignment = Alignment(
                horizontal="left",
                vertical="center",
                wrap_text=True,
            )

            for col_num in range(1, 5):
                ws.cell(row=row_num, column=col_num).border = label_border

            ws.row_dimensions[row_num].height = 38
            row_num += 1

    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_options.horizontalCentered = True

    margin = 0.25
    ws.page_margins.left = margin
    ws.page_margins.right = margin
    ws.page_margins.top = margin
    ws.page_margins.bottom = margin
    ws.page_margins.header = 0
    ws.page_margins.footer = 0

    wb.save(output_path)


def write_combined_easymark_output(output_path, sections):
    wb = Workbook()
    ws = wb.active
    ws.title = "EasyMark"
    ws.column_dimensions["A"].width = 70

    bold_font = InlineFont(b=True)
    normal_font = InlineFont()

    row_num = 1

    for section_index, section in enumerate(sections):
        item = section["item"]
        labels = section["labels"]

        if section_index > 0:
            row_num += 1

        header = ws.cell(row=row_num, column=1)
        header.value = make_section_header(item)
        header.font = Font(bold=True, size=11)
        header.alignment = Alignment(wrap_text=True, vertical="center")
        ws.row_dimensions[row_num].height = 24
        row_num += 1

        start_number = int(item.get("cable_start", 1))

        for index, label in enumerate(labels):
            cable_number = start_number + index

            path_value = label["path"]
            first_line = (
                f"{cable_number} - {path_value}"
                if path_value
                else str(cable_number)
            )

            project_identifier = item.get("project_number", "")
            identifier = item.get("identifier", "")

            if project_identifier and identifier:
                project_identifier = f"{project_identifier} - {identifier}"
            elif identifier:
                project_identifier = identifier

            rich_text = CellRichText(
                TextBlock(bold_font, first_line),
                "\n",
                TextBlock(bold_font, label["a_position"]),
                "\n",
                TextBlock(bold_font, label["z_position"]),
                "\n",
                TextBlock(normal_font, label["cable_type"]),
                "\n",
                TextBlock(normal_font, project_identifier),
                "\n",
                TextBlock(normal_font, item.get("other", "")),
            )

            cell = ws.cell(row=row_num, column=1)
            cell.value = rich_text
            cell.alignment = Alignment(
                horizontal="left",
                vertical="center",
                wrap_text=True,
            )

            ws.row_dimensions[row_num].height = 84
            row_num += 1

    ws.sheet_view.showGridLines = False
    wb.save(output_path)


# ----------------------------------------------------------------------
# Cable length calculation
# ----------------------------------------------------------------------

FIXED_FLOOR_SPACING_METERS = 0.35
FIXED_CORNER_ADD_METERS = 3.0
FIXED_SIDE_OFFSET_METERS = 0.7

DEFAULT_LENGTH_CONFIG = {
    "a_floor": 11,
    "b_floor": 94,
    "c_floor": 32,
    "d_floor": 70,
    "distance_to_wire_tray": 3.0,
    "row_pair_spacing": 2.44,
    "between_pair_spacing": 4.57,
    "romp_mode": "none",
    "romp_hallway_width": 4.57,
    "romp_row_interval": 40,
    "two_romp_spacing": 15.24,
    "two_romp_paired_row_spacing": 1.52,
    "two_romp_unpaired_row_spacing": 3.05,
}

COPPER_COLORS = (
    "GREEN",
    "BLUE",
    "RED",
    "YELLOW",
    "ORANGE",
    "PURPLE",
    "VIOLET",
    "WHITE",
    "BLACK",
    "GRAY",
    "GREY",
    "BROWN",
    "PINK",
    "AQUA",
    "TEAL",
)


def clone_length_config(building="Default"):
    config = dict(DEFAULT_LENGTH_CONFIG)
    if str(building) == "100":
        config["romp_mode"] = "building100TwoRomp"
    return config


def parse_bulk_location(value):
    """
    Accepts the old row-floor format (2-11) as well as longer spreadsheet
    position strings such as 01-01-002-83. The last two numeric groups are
    interpreted as row and floor/machine.
    """
    raw = clean_value(value)
    if not raw:
        return None

    numbers = re.findall(r"\d+", raw)
    if len(numbers) < 2:
        return None

    try:
        row = int(numbers[-2])
        machine = int(numbers[-1])
    except ValueError:
        return None

    if row < 1:
        return None

    return {
        "row": row,
        "machine": machine,
        "raw": raw,
    }


def infer_location_side(value):
    text = clean_value(value).lower()
    if re.search(r"\bright\b|\brhs\b", text):
        return "right"
    return "left"


def infer_romp(value):
    text = clean_value(value)
    match = re.search(r"\bromp\s*([12])\b", text, flags=re.I)
    if match:
        return match.group(1)

    match = re.search(r"\broom\s*([12])\b", text, flags=re.I)
    if match:
        return match.group(1)

    return "1"


def infer_building_from_text(*values):
    combined = " ".join(clean_value(value) for value in values)
    for building in ("100", "104", "105"):
        if re.search(rf"(?<!\d){building}(?!\d)", combined):
            return building
    return "Default"


def get_path_order(config):
    paths = [
        ("A", float(config["a_floor"])),
        ("B", float(config["b_floor"])),
        ("C", float(config["c_floor"])),
        ("D", float(config["d_floor"])),
    ]
    return "".join(path for path, _ in sorted(paths, key=lambda x: x[1]))


def get_path_side_pattern(path, config):
    order = get_path_order(config)
    try:
        index = order.index(path)
    except ValueError:
        return None
    return "first" if index < 2 else "second"


def get_side_offset_for_location(row, path, location_side, config):
    is_odd = row % 2 != 0
    pattern = get_path_side_pattern(path, config)

    if pattern == "first":
        if location_side == "left" and is_odd:
            return 0
        if location_side == "left" and not is_odd:
            return FIXED_SIDE_OFFSET_METERS
        if location_side == "right" and is_odd:
            return FIXED_SIDE_OFFSET_METERS
        if location_side == "right" and not is_odd:
            return 0

    if pattern == "second":
        if location_side == "left" and is_odd:
            return FIXED_SIDE_OFFSET_METERS
        if location_side == "left" and not is_odd:
            return 0
        if location_side == "right" and is_odd:
            return 0
        if location_side == "right" and not is_odd:
            return FIXED_SIDE_OFFSET_METERS

    return 0


def get_row_distance(row_a, row_b, config):
    start = min(row_a, row_b)
    end = max(row_a, row_b)
    distance = 0.0

    for row in range(start, end):
        if row % 2 != 0:
            distance += float(config["row_pair_spacing"])
        else:
            distance += float(config["between_pair_spacing"])

    return distance


def get_path_machine_number(path, config):
    return {
        "A": float(config["a_floor"]),
        "B": float(config["b_floor"]),
        "C": float(config["c_floor"]),
        "D": float(config["d_floor"]),
    }.get(path)


def get_floor_range(config):
    floors = [
        float(config["a_floor"]),
        float(config["b_floor"]),
        float(config["c_floor"]),
        float(config["d_floor"]),
    ]
    return min(floors), max(floors)


def feet_to_meters(value):
    return float(value) / 3.28084


def get_building100_special_row_distance(row_a, row_b, config):
    start = min(row_a, row_b)
    end = max(row_a, row_b)

    special_feet = {
        (5, 6): 12,
        (11, 12): 8,
        (12, 13): 14,
        (18, 19): 25,
        (24, 25): 20,
        (30, 31): 14,
        (34, 35): 36,
        (36, 37): 10,
        (38, 39): 20,
        (44, 45): 20,
        (48, 49): 12,
    }

    distance = 0.0

    for row in range(start, end):
        key = (row, row + 1)

        if key in special_feet:
            distance += feet_to_meters(special_feet[key])
        elif row >= 2 and row % 2 == 0:
            distance += float(config["two_romp_paired_row_spacing"])
        else:
            distance += float(config["two_romp_unpaired_row_spacing"])

    return distance


def get_building100_row_boundary_distance(start_row, boundary_row, config):
    distance = 0.0

    if start_row <= boundary_row:
        for row in range(start_row, boundary_row + 1):
            distance += get_building100_special_row_distance(
                row,
                row + 1,
                config,
            )
    else:
        for row in range(boundary_row + 1, start_row + 1):
            distance += get_building100_special_row_distance(
                row - 1,
                row,
                config,
            )

    return distance


def get_building100_path_d_crossing_boundary(row_a, row_b):
    highest = max(row_a, row_b)

    if highest <= 18:
        return 18
    if highest <= 36:
        return 36
    return 40


def get_building100_path_d_cross_romp_row_distance(row_a, row_b, config):
    boundary = get_building100_path_d_crossing_boundary(row_a, row_b)

    return (
        get_building100_row_boundary_distance(row_a, boundary, config)
        + get_building100_row_boundary_distance(row_b, boundary, config)
    )


def get_building100_full_romp_width(config):
    floor_min, floor_max = get_floor_range(config)
    return abs(floor_max - floor_min) * FIXED_FLOOR_SPACING_METERS


def get_romp_block(row, config):
    interval = max(1, int(config.get("romp_row_interval", 40) or 40))
    return (row - 1) // interval


def get_romp_adjustment(location_a, location_b, config):
    if config.get("romp_mode") != "evenOddEvery40":
        return 0.0

    hallway = float(config.get("romp_hallway_width", 0) or 0)
    crossings = 0

    if (location_a["row"] % 2) != (location_b["row"] % 2):
        crossings += 1

    crossings += abs(
        get_romp_block(location_a["row"], config)
        - get_romp_block(location_b["row"], config)
    )

    return crossings * hallway


def calculate_bulk_distance(
    location_a,
    location_b,
    path,
    config,
    location_a_side="left",
    location_b_side="left",
    room_a="1",
    room_b="1",
):
    path = clean_value(path).upper()[:1]
    path_machine = get_path_machine_number(path, config)

    if path_machine is None:
        return None

    floor_min, floor_max = get_floor_range(config)

    if (
        location_a["machine"] < floor_min
        or location_a["machine"] > floor_max
        or location_b["machine"] < floor_min
        or location_b["machine"] > floor_max
    ):
        return "INVALID_FLOOR"

    different_100_room = (
        config.get("romp_mode") == "building100TwoRomp"
        and str(room_a) != str(room_b)
    )

    distance = 0.0

    if location_a["row"] == location_b["row"] and not different_100_room:
        distance += (
            abs(location_a["machine"] - location_b["machine"])
            * FIXED_FLOOR_SPACING_METERS
        )
    else:
        start_to_path = abs(location_a["machine"] - path_machine)
        path_to_destination = abs(location_b["machine"] - path_machine)
        room_crossing_distance = 0.0

        if config.get("romp_mode") == "building100TwoRomp" and different_100_room:
            start_connector = 94 if str(room_a) == "1" else 11
            destination_connector = 94 if str(room_b) == "1" else 11

            start_to_path = abs(location_a["machine"] - start_connector)
            path_to_destination = abs(
                location_b["machine"] - destination_connector
            )
            room_crossing_distance = float(
                config.get("two_romp_spacing", 15.24)
            )

        machine_distance = (
            start_to_path + path_to_destination
        ) * FIXED_FLOOR_SPACING_METERS

        if (
            config.get("romp_mode") == "building100TwoRomp"
            and path == "D"
            and different_100_room
        ):
            row_distance = get_building100_path_d_cross_romp_row_distance(
                location_a["row"],
                location_b["row"],
                config,
            )
        elif config.get("romp_mode") == "building100TwoRomp":
            row_distance = get_building100_special_row_distance(
                location_a["row"],
                location_b["row"],
                config,
            )
        else:
            row_distance = get_row_distance(
                location_a["row"],
                location_b["row"],
                config,
            )

        distance += machine_distance
        distance += row_distance
        distance += room_crossing_distance
        distance += get_romp_adjustment(
            location_a,
            location_b,
            config,
        )
        distance += FIXED_CORNER_ADD_METERS

        if (
            config.get("romp_mode") == "building100TwoRomp"
            and path == "D"
            and different_100_room
        ):
            distance += get_building100_full_romp_width(config)

    distance += get_side_offset_for_location(
        location_a["row"],
        path,
        location_a_side,
        config,
    )
    distance += get_side_offset_for_location(
        location_b["row"],
        path,
        location_b_side,
        config,
    )
    distance += float(config["distance_to_wire_tray"]) * 2

    return distance


def round_up_to_nearest_five(value):
    if value is None:
        return None
    return int((float(value) + 4.999999999) // 5) * 5


def meters_to_whole_feet(value):
    if value is None:
        return None
    return int(math.ceil(float(value) * 3.280839895))


def extract_path_letter(value):
    text = clean_value(value).upper()

    match = re.search(r"(?:^|[^A-Z])([ABCD])(?:$|[^A-Z])", text)
    if match:
        return match.group(1)

    if text in ("A", "B", "C", "D"):
        return text

    return ""


def choose_shortest_path(
    location_a,
    location_b,
    config,
    side_a,
    side_b,
    room_a,
    room_b,
):
    options = []

    for path in ("A", "B", "C", "D"):
        distance = calculate_bulk_distance(
            location_a,
            location_b,
            path,
            config,
            side_a,
            side_b,
            room_a,
            room_b,
        )

        if isinstance(distance, (int, float)):
            options.append((distance, path))

    if not options:
        return "", None

    distance, path = min(options)
    return path, distance


def copper_color_from_type(cable_type):
    text = clean_value(cable_type).upper()

    for color in COPPER_COLORS:
        if re.search(rf"\b{re.escape(color)}\b", text):
            return "GRAY" if color == "GREY" else color

    before = text.split("COPPER", 1)[0].strip(" -_/")
    if before:
        return before

    return "UNSPECIFIED"


def cable_length_category(cable_type):
    text = clean_value(cable_type).upper()

    if "COPPER" in text:
        return "Copper"

    if re.search(r"\b\d+F\b", text):
        return "Fiber"

    return "Other"


def cable_length_subtype(cable_type):
    text = clean_value(cable_type).upper()

    if "COPPER" in text:
        color = copper_color_from_type(text)

        if color and color != "UNSPECIFIED":
            return f"{color.title()} Copper"

        return "Copper"

    fiber_match = re.search(r"\b(\d+F)\b", text)
    if fiber_match:
        return fiber_match.group(1).upper()

    return "Other"


def cable_matches_length_filter(cable_type, filter_name):
    category = cable_length_category(cable_type)
    subtype = cable_length_subtype(cable_type)

    if filter_name == "All":
        return True

    if filter_name == "Fiber":
        return category == "Fiber"

    if filter_name == "Copper":
        return category == "Copper"

    if filter_name == "Other":
        return category == "Other"

    return subtype == filter_name


def summarize_cable_lengths(records):
    non_copper_grouped = {}
    copper_grouped = {}
    copper_totals = {}

    for record in records:
        if record.get("status") != "OK":
            continue

        cable_type = record["cable_type"]
        raw_m = float(record["raw_length_m"])
        is_copper = "COPPER" in cable_type.upper()

        if is_copper:
            color = copper_color_from_type(cable_type)
            length_ft = meters_to_whole_feet(raw_m)
            key = (cable_type, length_ft)

            row = copper_grouped.setdefault(
                key,
                {
                    "cable_type": cable_type,
                    "length": f"{length_ft} ft",
                    "quantity": 0,
                    "total_length": 0,
                    "is_copper": True,
                },
            )
            row["quantity"] += 1
            row["total_length"] += length_ft
            copper_totals[color] = copper_totals.get(color, 0) + length_ft

        else:
            rounded_m = round_up_to_nearest_five(raw_m)
            key = (cable_type, rounded_m)

            row = non_copper_grouped.setdefault(
                key,
                {
                    "cable_type": cable_type,
                    "length": f"{rounded_m} m",
                    "quantity": 0,
                    "total_length": "",
                    "is_copper": False,
                },
            )
            row["quantity"] += 1

    non_copper_rows = list(non_copper_grouped.values())
    non_copper_rows.sort(
        key=lambda item: (
            item["cable_type"].upper(),
            item["length"],
        )
    )

    copper_rows = list(copper_grouped.values())
    copper_rows.sort(
        key=lambda item: (
            copper_color_from_type(item["cable_type"]),
            item["cable_type"].upper(),
            int(str(item["length"]).split()[0]),
        )
    )

    rows = list(non_copper_rows)

    # Visually separate fiber/other from copper.
    if non_copper_rows and (copper_rows or copper_totals):
        rows.append(
            {
                "cable_type": "",
                "length": "",
                "quantity": "",
                "total_length": "",
                "separator": True,
                "is_copper": False,
            }
        )

    # All individual copper lengths stay below all non-copper rows.
    rows.extend(copper_rows)

    # Copper color totals are always the final rows.
    for color, total_ft in sorted(copper_totals.items()):
        rows.append(
            {
                "cable_type": f"{color} COPPER TOTAL",
                "length": "",
                "quantity": "",
                "total_length": f"{int(total_ft)} ft",
                "is_copper": True,
            }
        )

    return rows


def write_cable_length_output(output_path, summary_rows, raw_records):
    wb = Workbook()

    summary_ws = wb.active
    summary_ws.title = "Cable List"
    summary_ws.append(
        [
            "Cable Type",
            "Length",
            "Quantity",
            "Total Length",
        ]
    )

    for row in summary_rows:
        if row.get("separator"):
            summary_ws.append(["", "", "", ""])
            continue

        summary_ws.append(
            [
                row["cable_type"],
                row["length"],
                row["quantity"],
                row["total_length"],
            ]
        )

    raw_ws = wb.create_sheet("Raw Cable Lengths")
    raw_ws.append(
        [
            "File",
            "Worksheet",
            "Building",
            "A Position",
            "Z Position",
            "Path",
            "Cable Type",
            "Raw Length",
            "Rounded Length",
            "Status",
        ]
    )

    for record in raw_records:
        is_copper = "COPPER" in record["cable_type"].upper()

        if is_copper and isinstance(record.get("raw_length_m"), (int, float)):
            raw_display = f'{meters_to_whole_feet(record["raw_length_m"])} ft'
            rounded_display = raw_display
        else:
            raw_display = (
                f'{round(record["raw_length_m"], 2)} m'
                if isinstance(record.get("raw_length_m"), (int, float))
                else ""
            )
            rounded_display = (
                f'{record["rounded_length_m"]} m'
                if record.get("rounded_length_m") is not None
                else ""
            )

        raw_ws.append(
            [
                record["file_name"],
                record["sheet"],
                record["building"],
                record["a_position"],
                record["z_position"],
                record["path"],
                record["cable_type"],
                raw_display,
                rounded_display,
                record["status"],
            ]
        )

    for ws in (summary_ws, raw_ws):
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        for column_cells in ws.columns:
            max_len = 0
            letter = get_column_letter(column_cells[0].column)

            for cell in column_cells:
                max_len = max(max_len, len(clean_value(cell.value)))

            ws.column_dimensions[letter].width = min(
                max(max_len + 2, 12),
                45,
            )

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    wb.save(output_path)


def set_windows_app_identity():
    """Give the packaged app a stable Windows taskbar identity."""
    if not sys.platform.startswith("win"):
        return

    try:
        app_id = "MTG.MTGCableTool"
        ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID(
            app_id
        )
    except Exception:
        # Do not prevent the program from opening if Windows does not
        # support the API or the call fails for any reason.
        pass


class LabelGeneratorApp:
    def __init__(self, root):
        self.root = root
        self.root.title(APP_NAME)
        self.root.geometry("1100x760")
        self.root.minsize(700, 500)
        self.root.resizable(True, True)

        self.file_path = tk.StringVar()
        self.sheet_name = tk.StringVar()

        self.a_column = tk.StringVar()
        self.z_column = tk.StringVar()
        self.cable_type_column = tk.StringVar()
        self.path_column = tk.StringVar()

        self.current_header_row = None
        self.current_auto_columns = {}

        # Shared spreadsheet caches used by all three tabs.
        # Cache entries automatically invalidate when a source file's
        # modification time or size changes.
        self.workbook_cache = {}
        self.label_cache = {}

        # Batch / multi-file workflow state.
        self.batch_items = []
        self.batch_file_paths = []

        self.batch_project_number = tk.StringVar()
        self.batch_cable_start = tk.StringVar(value="1")
        self.batch_identifier = tk.StringVar()
        self.batch_other = tk.StringVar()
        self.batch_selected_index = None

        self.output_type = tk.StringVar(value="Labels")
        self.batch_output_type = tk.StringVar(value="Labels")

        self.batch_a_column = tk.StringVar()
        self.batch_z_column = tk.StringVar()
        self.batch_cable_type_column = tk.StringVar()
        self.batch_path_column = tk.StringVar()
        self.batch_mapping_combos = {}
        self.batch_cell_editor = None
        self.batch_cell_editor_info = None

        self.loading_text = tk.StringVar(value="Ready")

        # Cable-length workflow state.
        self.length_items = []
        self.length_file_paths = []
        self.length_selected_index = None
        self.length_building = tk.StringVar(value="Default")
        self.length_filter = tk.StringVar(value="All")

        self.length_a_column = tk.StringVar()
        self.length_z_column = tk.StringVar()
        self.length_cable_type_column = tk.StringVar()
        self.length_path_column = tk.StringVar()
        self.length_mapping_combos = {}

        self.length_config_building = tk.StringVar(value="Default")
        self.length_a_floor = tk.StringVar(value="11")
        self.length_b_floor = tk.StringVar(value="94")
        self.length_c_floor = tk.StringVar(value="32")
        self.length_d_floor = tk.StringVar(value="70")
        self.length_wire_tray = tk.StringVar(value="3")
        self.length_paired_rows = tk.StringVar(value="2.44")
        self.length_aisles = tk.StringVar(value="4.57")

        self.length_configs = {
            building: clone_length_config(building)
            for building in ("100", "104", "105")
        }
        self.length_configs["Default"] = dict(self.length_configs["105"])
        self.length_building_editor_visible = False

        self.length_raw_records = []
        self.length_summary_rows = []

        self.project_number = tk.StringVar()
        self.starting_cable_number = tk.StringVar(value="1")
        self.identifier = tk.StringVar()
        self.other = tk.StringVar()
        self.status_text = tk.StringVar(value="Select an Excel file to begin.")

        self.build_ui()

    def file_cache_signature(self, file_path):
        try:
            stat = os.stat(file_path)
            return (
                stat.st_mtime_ns,
                stat.st_size,
            )
        except OSError:
            return None

    def normalized_cache_path(self, file_path):
        return os.path.normcase(
            os.path.abspath(file_path)
        )

    def invalidate_file_cache_if_changed(self, file_path):
        normalized = self.normalized_cache_path(file_path)
        signature = self.file_cache_signature(file_path)

        cached = self.workbook_cache.get(normalized)

        if cached and cached.get("signature") != signature:
            self.workbook_cache.pop(normalized, None)

            stale_keys = [
                key
                for key in self.label_cache
                if key[0] == normalized
            ]

            for key in stale_keys:
                self.label_cache.pop(key, None)

        return normalized, signature

    def cache_sheet_from_open_worksheet(
        self,
        file_path,
        ws,
        header_row=None,
        columns=None,
        include_preview=False,
    ):
        normalized, signature = self.invalidate_file_cache_if_changed(
            file_path
        )

        workbook_entry = self.workbook_cache.setdefault(
            normalized,
            {
                "signature": signature,
                "sheet_names": [],
                "sheets": {},
            },
        )
        workbook_entry["signature"] = signature

        if ws.title not in workbook_entry["sheet_names"]:
            workbook_entry["sheet_names"].append(ws.title)

        if header_row is None or columns is None:
            header_row, columns = find_header_columns(ws)

        sheet_entry = workbook_entry["sheets"].setdefault(
            ws.title,
            {},
        )

        sheet_entry["header_row"] = header_row
        sheet_entry["auto_columns"] = dict(columns or {})
        sheet_entry["max_column"] = ws.max_column
        sheet_entry["max_row"] = ws.max_row

        if header_row:
            choices = []
            choice_by_col = {}

            for col_num in range(1, ws.max_column + 1):
                choice = self.column_choice(
                    col_num,
                    ws.cell(
                        row=header_row,
                        column=col_num,
                    ).value,
                )
                choices.append(choice)
                choice_by_col[col_num] = choice

            sheet_entry["choices"] = choices
            sheet_entry["choice_by_col"] = choice_by_col

            # Pre-cache the union of source columns used by all tabs.
            shared_columns = {
                key: columns.get(key)
                for key in SHARED_SOURCE_FIELDS
                if columns.get(key)
            }

            sheet_entry["shared_columns"] = dict(shared_columns)

            if all(
                shared_columns.get(key)
                for key in (
                    "a_position",
                    "z_position",
                    "cable_type",
                )
            ):
                raw_rows = []
                data_started = False
                blank_run = 0

                for row_num in range(
                    header_row + 1,
                    ws.max_row + 1,
                ):
                    row_values = {}

                    for key, col_num in shared_columns.items():
                        row_values[key] = clean_value(
                            ws.cell(
                                row=row_num,
                                column=col_num,
                            ).value
                        )

                    row_is_blank = not any(
                        row_values.get(key, "")
                        for key in SHARED_SOURCE_FIELDS
                    )

                    if row_is_blank:
                        if data_started:
                            blank_run += 1

                            if blank_run >= EXTRACT_BLANK_ROW_STOP:
                                break

                        continue

                    data_started = True
                    blank_run = 0

                    raw_rows.append(
                        {
                            "row_num": row_num,
                            "values": row_values,
                        }
                    )

                sheet_entry["shared_raw_rows"] = raw_rows

        return sheet_entry


    def get_cached_workbook_overview(
        self,
        file_path,
        include_preview_sheet=None,
    ):
        normalized, signature = self.invalidate_file_cache_if_changed(
            file_path
        )

        cached = self.workbook_cache.get(normalized)

        if (
            cached
            and cached.get("signature") == signature
            and cached.get("overview_complete")
        ):
            return cached

        workbook = load_workbook(
            file_path,
            data_only=True,
            read_only=True,
        )

        try:
            self.workbook_cache[normalized] = {
                "signature": signature,
                "sheet_names": list(workbook.sheetnames),
                "sheets": {},
                "overview_complete": False,
            }

            for ws in workbook.worksheets:
                self.cache_sheet_from_open_worksheet(
                    file_path,
                    ws,
                )

            entry = self.workbook_cache[normalized]
            entry["overview_complete"] = True
            return entry

        finally:
            workbook.close()

    def get_cached_sheet_info(
        self,
        file_path,
        sheet_name,
        include_preview=False,
    ):
        normalized, signature = self.invalidate_file_cache_if_changed(
            file_path
        )

        cached = self.workbook_cache.get(normalized)
        sheet_entry = None

        if cached and cached.get("signature") == signature:
            sheet_entry = cached.get("sheets", {}).get(sheet_name)

            if sheet_entry:
                return sheet_entry

        workbook = load_workbook(
            file_path,
            data_only=True,
            read_only=True,
        )

        try:
            if sheet_name not in workbook.sheetnames:
                raise RuntimeError(
                    f'Worksheet "{sheet_name}" was not found.'
                )

            ws = workbook[sheet_name]

            return self.cache_sheet_from_open_worksheet(
                file_path,
                ws,
                include_preview=include_preview,
            )

        finally:
            workbook.close()

    def cached_extract_labels(
        self,
        file_path,
        sheet_name=None,
        header_row_override=None,
        columns_override=None,
        progress_callback=None,
    ):
        normalized, signature = self.invalidate_file_cache_if_changed(
            file_path
        )

        column_key = tuple(
            sorted(
                (columns_override or {}).items()
            )
        )

        key = (
            normalized,
            sheet_name or "",
            header_row_override,
            column_key,
            signature,
        )

        cached = self.label_cache.get(key)

        if cached is not None:
            return [
                dict(label)
                for label in cached
            ]

        # Fast path: if this request uses the sheet's detected shared
        # A/Z/Cable Type/Path mapping, build the result entirely from RAM.
        labels = None

        if sheet_name and columns_override:
            try:
                info = self.get_cached_sheet_info(
                    file_path,
                    sheet_name,
                )

                shared_columns = dict(
                    info.get("shared_columns", {})
                )
                shared_rows = info.get("shared_raw_rows")

                requested_columns = {
                    key: value
                    for key, value in columns_override.items()
                    if value
                }

                mapping_matches_cache = all(
                    shared_columns.get(key) == col_num
                    for key, col_num in requested_columns.items()
                )

                required_present = all(
                    requested_columns.get(key)
                    for key in (
                        "a_position",
                        "z_position",
                        "cable_type",
                    )
                )

                header_matches = (
                    header_row_override is None
                    or header_row_override == info.get("header_row")
                )

                if (
                    shared_rows is not None
                    and mapping_matches_cache
                    and required_present
                    and header_matches
                ):
                    labels = []

                    for raw in shared_rows:
                        values = raw["values"]
                        cable_type = clean_value(
                            values.get("cable_type", "")
                        )

                        # Blank Cable Type means continuation row.
                        if not cable_type:
                            continue

                        labels.append(
                            {
                                "a_position": clean_value(
                                    values.get("a_position", "")
                                ),
                                "z_position": clean_value(
                                    values.get("z_position", "")
                                ),
                                "cable_type": simplify_cable_type(
                                    cable_type
                                ),
                                "path": clean_value(
                                    values.get("path", "")
                                ),
                                "source_sheet": sheet_name,
                                "source_row": raw["row_num"],
                            }
                        )

            except Exception:
                labels = None

        # Fallback for manual mappings or any request not covered by the
        # shared detected-column cache.
        if labels is None:
            labels = extract_labels(
                file_path,
                sheet_name,
                header_row_override=header_row_override,
                columns_override=columns_override,
                progress_callback=progress_callback,
            )

        self.label_cache[key] = [
            dict(label)
            for label in labels
        ]

        return [
            dict(label)
            for label in labels
        ]


    def sanitize_filename_part(self, value):
        value = clean_value(value).strip()

        if not value:
            return ""

        # Windows-safe filename characters.
        value = re.sub(r'[<>:"/\\|?*]+', "-", value)
        value = re.sub(r"\s+", " ", value).strip()
        value = value.strip(". ")

        return value

    def find_batch_metadata_for_sheet(self, item):
        file_path = os.path.normcase(
            os.path.abspath(item.get("file_path", ""))
        )
        sheet = clean_value(item.get("sheet", ""))

        for batch_item in self.batch_items:
            batch_path = os.path.normcase(
                os.path.abspath(batch_item.get("file_path", ""))
            )

            if (
                batch_path == file_path
                and clean_value(batch_item.get("sheet", "")) == sheet
            ):
                return batch_item

        return item

    def metadata_output_filename(self, item):
        if item is None:
            metadata = {}
        else:
            metadata = self.find_batch_metadata_for_sheet(item)

        parts = [
            self.sanitize_filename_part(
                metadata.get("project_number", "")
            ),
            self.sanitize_filename_part(
                metadata.get("identifier", "")
            ),
            self.sanitize_filename_part(
                metadata.get("other", "")
            ),
            date.today().strftime("%Y-%m-%d"),
        ]

        parts = [
            part
            for part in parts
            if part
        ]

        return "_".join(parts) + ".xlsx"

    def first_batch_output_filename(self, output_type="Labels"):
        item = self.batch_items[0] if self.batch_items else None
        base_name = self.metadata_output_filename(item)

        if base_name.lower().endswith(".xlsx"):
            base_name = base_name[:-5]

        suffixes = {
            "Labels": "Labels",
            "Printable Labels": "Printable",
            "Easy-Mark": "EasyMark",
        }

        suffix = suffixes.get(
            output_type,
            self.sanitize_filename_part(output_type) or "Labels",
        )

        return f"{base_name}_{suffix}.xlsx"

    def first_length_output_filename(self):
        if self.length_items:
            first_sheet = self.sanitize_filename_part(
                self.length_items[0].get("sheet", "")
            )
        else:
            first_sheet = ""

        if not first_sheet:
            first_sheet = "Cable_Lengths"

        return f"{first_sheet}_Lengths.xlsx"

    def build_ui(self):
        outer = ttk.Frame(
            self.root,
            padding=12,
        )
        outer.pack(
            fill="both",
            expand=True,
        )

        title_row = ttk.Frame(outer)
        title_row.pack(
            fill="x",
            pady=(0, 10),
        )

        ttk.Label(
            title_row,
            text="MTG Cable Tool",
            font=("TkDefaultFont", 16, "bold"),
        ).pack(side="left")

        self.version_label = ttk.Label(
            title_row,
            text=f"v{APP_VERSION}",
        )
        self.version_label.pack(
            side="left",
            padx=(10, 0),
        )

        # Hidden when idle.
        self.loading_label = ttk.Label(
            title_row,
            text="Loading...",
            font=("TkDefaultFont", 10, "bold"),
        )

        self.main_notebook = ttk.Notebook(outer)
        self.main_notebook.pack(
            fill="both",
            expand=True,
        )

        # The notebook tabs stay fixed. Each tab gets its own scrollable
        # content area underneath the tab bar.
        self.single_tab = self.create_scrollable_tab(
            self.main_notebook,
            "Single File",
        )
        self.batch_tab = self.create_scrollable_tab(
            self.main_notebook,
            "Multiple Files / Sheets",
        )
        self.length_tab = self.create_scrollable_tab(
            self.main_notebook,
            "Cable Lengths",
        )

        self.build_single_tab()
        self.build_batch_tab()
        self.build_length_tab()

        self.status_text = tk.StringVar(
            value="Select an Excel file to begin."
        )

        status = ttk.Label(
            outer,
            textvariable=self.status_text,
            relief="sunken",
            anchor="w",
            padding=8,
        )
        status.pack(
            fill="x",
            pady=(10, 0),
        )

    def create_scrollable_tab(self, notebook, title):
        holder = ttk.Frame(notebook)
        notebook.add(holder, text=title)

        holder.rowconfigure(0, weight=1)
        holder.columnconfigure(0, weight=1)

        canvas = tk.Canvas(
            holder,
            highlightthickness=0,
        )
        canvas.grid(
            row=0,
            column=0,
            sticky="nsew",
        )

        vscroll = ttk.Scrollbar(
            holder,
            orient="vertical",
            command=canvas.yview,
        )
        vscroll.grid(
            row=0,
            column=1,
            sticky="ns",
        )

        hscroll = ttk.Scrollbar(
            holder,
            orient="horizontal",
            command=canvas.xview,
        )
        hscroll.grid(
            row=1,
            column=0,
            sticky="ew",
        )

        canvas.configure(
            yscrollcommand=vscroll.set,
            xscrollcommand=hscroll.set,
        )

        content = ttk.Frame(
            canvas,
            padding=10,
        )

        window_id = canvas.create_window(
            (0, 0),
            window=content,
            anchor="nw",
        )

        def update_scroll_region(event=None):
            canvas.configure(
                scrollregion=canvas.bbox("all")
            )

        def resize_content(event):
            requested_width = content.winfo_reqwidth()
            width = max(
                event.width,
                requested_width,
            )
            canvas.itemconfigure(
                window_id,
                width=width,
            )

        content.bind(
            "<Configure>",
            update_scroll_region,
        )
        canvas.bind(
            "<Configure>",
            resize_content,
        )

        def on_mousewheel(event):
            if event.delta:
                canvas.yview_scroll(
                    int(-1 * (event.delta / 120)),
                    "units",
                )

        def on_shift_mousewheel(event):
            if event.delta:
                canvas.xview_scroll(
                    int(-1 * (event.delta / 120)),
                    "units",
                )

        # Only scroll the active tab when the pointer is over its canvas.
        canvas.bind(
            "<MouseWheel>",
            on_mousewheel,
        )
        canvas.bind(
            "<Shift-MouseWheel>",
            on_shift_mousewheel,
        )

        return content


    def start_loading(self, message="Loading..."):
        self.loading_text.set(message)
        self.loading_label.configure(text="Loading...")

        if not self.loading_label.winfo_ismapped():
            self.loading_label.pack(
                side="left",
                padx=(12, 0),
            )

        # Paint Loading... immediately before the program starts the task.
        self.root.update_idletasks()
        self.root.update()

    def pulse_loading(self):
        # Existing scan callbacks can keep calling this, but no animation
        # or progress-bar work is performed.
        return

    def stop_loading(self):
        if self.loading_label.winfo_ismapped():
            self.loading_label.pack_forget()

        self.loading_text.set("Ready")
        self.root.update_idletasks()


    def build_single_tab(self):
        main = self.single_tab

        top_area = ttk.Frame(main)
        top_area.pack(fill="x")

        file_frame = ttk.LabelFrame(top_area, text="Source Excel File", padding=10)
        file_frame.pack(fill="x", pady=(0, 8))

        file_entry = ttk.Entry(file_frame, textvariable=self.file_path)
        file_entry.pack(side="left", fill="x", expand=True, padx=(0, 10))

        ttk.Button(
            file_frame,
            text="Select Excel File",
            command=self.select_file,
        ).pack(side="right")

        sheet_frame = ttk.Frame(top_area)
        sheet_frame.pack(fill="x", pady=(0, 8))

        ttk.Label(sheet_frame, text="Worksheet:").pack(side="left", padx=(0, 10))

        self.sheet_combo = ttk.Combobox(
            sheet_frame,
            textvariable=self.sheet_name,
            state="readonly",
        )
        self.sheet_combo.pack(side="left", fill="x", expand=True)
        self.sheet_combo.bind("<<ComboboxSelected>>", self.on_sheet_selected)

        middle = ttk.Frame(main)
        middle.pack(fill="x", pady=(0, 8))

        fields_frame = ttk.LabelFrame(
            middle,
            text="Label Information",
            padding=10,
        )
        fields_frame.pack(
            side="left",
            fill="both",
            expand=True,
            padx=(0, 6),
        )

        fields = [
            ("Project Number", self.project_number),
            ("Starting Cable Number", self.starting_cable_number),
            ("Identifier", self.identifier),
            ("Other", self.other),
        ]

        for row_num, (label_text, variable) in enumerate(fields):
            ttk.Label(
                fields_frame,
                text=label_text + ":",
            ).grid(
                row=row_num,
                column=0,
                sticky="w",
                padx=(0, 10),
                pady=3,
            )
            ttk.Entry(
                fields_frame,
                textvariable=variable,
            ).grid(
                row=row_num,
                column=1,
                sticky="ew",
                pady=3,
            )

        fields_frame.columnconfigure(1, weight=1)

        mapping_frame = ttk.LabelFrame(
            middle,
            text="Column Mapping",
            padding=10,
        )
        mapping_frame.pack(
            side="left",
            fill="both",
            expand=True,
            padx=(6, 0),
        )

        mapping_fields = [
            ("A Position", self.a_column, "a_position"),
            ("Z Position", self.z_column, "z_position"),
            ("Cable Type", self.cable_type_column, "cable_type"),
            ("Path (optional)", self.path_column, "path"),
        ]

        self.mapping_combos = {}

        for row_num, (label_text, variable, key) in enumerate(mapping_fields):
            ttk.Label(
                mapping_frame,
                text=label_text + ":",
            ).grid(
                row=row_num,
                column=0,
                sticky="w",
                padx=(0, 10),
                pady=3,
            )

            combo = ttk.Combobox(
                mapping_frame,
                textvariable=variable,
                state="readonly",
            )
            combo.grid(
                row=row_num,
                column=1,
                sticky="ew",
                pady=3,
            )
            combo.bind(
                "<<ComboboxSelected>>",
                self.on_mapping_changed,
            )
            self.mapping_combos[key] = combo

        mapping_frame.columnconfigure(1, weight=1)

        output_controls = ttk.Frame(main)
        output_controls.pack(fill="x", pady=(0, 8))

        ttk.Label(output_controls, text="Output Type:").pack(
            side="left",
            padx=(0, 8),
        )

        self.output_type_combo = ttk.Combobox(
            output_controls,
            textvariable=self.output_type,
            values=("Labels", "Printable Labels", "Easy-Mark"),
            state="readonly",
            width=20,
        )
        self.output_type_combo.pack(side="left")
        self.output_type_combo.bind(
            "<<ComboboxSelected>>",
            self.on_output_type_changed,
        )

        ttk.Button(
            output_controls,
            text="Generate",
            command=self.generate_selected_output,
        ).pack(side="left", padx=(8, 0))

        ttk.Button(
            output_controls,
            text="Refresh Preview",
            command=self.refresh_previews,
        ).pack(side="left", padx=(8, 0))

        ttk.Button(
            output_controls,
            text="Clear",
            command=self.clear_fields,
        ).pack(side="right")

        output_preview_frame = ttk.LabelFrame(
            main,
            text="Output Preview",
            padding=8,
        )
        output_preview_frame.pack(
            fill="both",
            expand=True,
        )

        self.output_preview = ttk.Treeview(
            output_preview_frame,
            show="headings",
        )
        output_v = ttk.Scrollbar(
            output_preview_frame,
            orient="vertical",
            command=self.output_preview.yview,
        )
        output_h = ttk.Scrollbar(
            output_preview_frame,
            orient="horizontal",
            command=self.output_preview.xview,
        )
        self.output_preview.configure(
            yscrollcommand=output_v.set,
            xscrollcommand=output_h.set,
        )
        self.output_preview.grid(row=0, column=0, sticky="nsew")
        output_v.grid(row=0, column=1, sticky="ns")
        output_h.grid(row=1, column=0, sticky="ew")
        output_preview_frame.rowconfigure(0, weight=1)
        output_preview_frame.columnconfigure(0, weight=1)


    def build_batch_tab(self):
        main = self.batch_tab

        controls = ttk.Frame(main)
        controls.pack(fill="x", pady=(0, 8))

        ttk.Button(
            controls,
            text="Add Excel Files",
            command=self.batch_add_files,
        ).pack(side="left")

        ttk.Button(
            controls,
            text="Remove Selected",
            command=self.batch_remove_selected,
        ).pack(side="left", padx=(8, 0))

        ttk.Button(
            controls,
            text="Clear List",
            command=self.batch_clear,
        ).pack(side="left", padx=(8, 0))

        ttk.Label(
            controls,
            text="Output Type:",
        ).pack(side="left", padx=(24, 8))

        self.batch_output_type_combo = ttk.Combobox(
            controls,
            textvariable=self.batch_output_type,
            values=("Labels", "Printable Labels", "Easy-Mark"),
            state="readonly",
            width=20,
        )
        self.batch_output_type_combo.pack(side="left")
        self.batch_output_type_combo.bind(
            "<<ComboboxSelected>>",
            self.on_batch_output_type_changed,
        )

        ttk.Button(
            controls,
            text="Generate",
            command=self.generate_selected_batch_output,
        ).pack(side="left", padx=(8, 0))

        list_frame = ttk.LabelFrame(
            main,
            text="Files and Worksheets",
            padding=8,
        )
        list_frame.pack(fill="both", expand=True, pady=(0, 8))

        columns = (
            "include",
            "file",
            "sheet",
            "project",
            "start",
            "identifier",
            "other",
            "a",
            "z",
            "type",
            "path",
            "count",
        )

        self.batch_tree = ttk.Treeview(
            list_frame,
            columns=columns,
            show="headings",
            selectmode="extended",
        )

        headings = {
            "include": "USE",
            "file": "File",
            "sheet": "Worksheet",
            "project": "Project Number",
            "start": "Cable Start",
            "identifier": "Identifier",
            "other": "Other",
            "a": "A Pos",
            "z": "Z Pos",
            "type": "Cable Type",
            "path": "Path",
            "count": "Cables",
        }

        for key in columns:
            self.batch_tree.heading(key, text=headings[key])
            self.batch_tree.column(
                key,
                width=90,
                anchor="center",
                stretch=True,
            )

        self.batch_tree.column(
            "include",
            width=90,
            stretch=False,
        )
        self.batch_tree.column("file", anchor="w")
        self.batch_tree.column("sheet", anchor="w")
        self.batch_tree.column("project", anchor="w")
        self.batch_tree.column("identifier", anchor="w")
        self.batch_tree.column("other", anchor="w")

        # Bold font for the whole batch table so the USE state is obvious.
        batch_style = ttk.Style()
        batch_style.configure(
            "Batch.Treeview",
            font=("TkDefaultFont", 10, "bold"),
            rowheight=26,
        )
        self.batch_tree.configure(style="Batch.Treeview")

        batch_v = ttk.Scrollbar(
            list_frame,
            orient="vertical",
            command=self.batch_tree.yview,
        )
        batch_h = ttk.Scrollbar(
            list_frame,
            orient="horizontal",
            command=self.batch_tree.xview,
        )
        self.batch_tree.configure(
            yscrollcommand=batch_v.set,
            xscrollcommand=batch_h.set,
        )

        self.batch_tree.grid(
            row=0,
            column=0,
            sticky="nsew",
        )
        batch_v.grid(row=0, column=1, sticky="ns")
        batch_h.grid(row=1, column=0, sticky="ew")
        list_frame.rowconfigure(0, weight=1)
        list_frame.columnconfigure(0, weight=1)

        self.batch_tree.bind(
            "<Button-1>",
            self.batch_tree_click,
            add="+",
        )
        self.batch_tree.bind(
            "<<TreeviewSelect>>",
            self.batch_load_selected_settings,
        )

        editor = ttk.Frame(main)
        editor.pack(fill="x")

        settings_frame = ttk.LabelFrame(
            editor,
            text="Selected Worksheet Settings",
            padding=10,
        )
        settings_frame.pack(
            side="left",
            fill="both",
            expand=True,
            padx=(0, 5),
        )

        setting_fields = [
            ("Project Number", self.batch_project_number),
            ("Cable Start", self.batch_cable_start),
            ("Identifier", self.batch_identifier),
            ("Other", self.batch_other),
        ]

        for row_num, (label_text, variable) in enumerate(setting_fields):
            ttk.Label(
                settings_frame,
                text=label_text + ":",
            ).grid(
                row=row_num,
                column=0,
                sticky="w",
                padx=(0, 8),
                pady=3,
            )
            ttk.Entry(
                settings_frame,
                textvariable=variable,
            ).grid(
                row=row_num,
                column=1,
                sticky="ew",
                pady=3,
            )

        settings_frame.columnconfigure(1, weight=1)

        settings_buttons = ttk.Frame(settings_frame)
        settings_buttons.grid(
            row=len(setting_fields),
            column=0,
            columnspan=2,
            sticky="e",
            pady=(8, 0),
        )

        ttk.Button(
            settings_buttons,
            text="Apply to Selected",
            command=self.batch_apply_settings,
        ).pack(side="left")

        ttk.Button(
            settings_buttons,
            text="Apply to All",
            command=self.batch_apply_settings_all,
        ).pack(side="left", padx=(8, 0))

        mapping_frame = ttk.LabelFrame(
            editor,
            text="Selected Worksheet Column Mapping",
            padding=10,
        )
        mapping_frame.pack(
            side="left",
            fill="both",
            expand=True,
            padx=(5, 0),
        )

        mapping_fields = [
            ("A Position", self.batch_a_column, "a_position"),
            ("Z Position", self.batch_z_column, "z_position"),
            ("Cable Type", self.batch_cable_type_column, "cable_type"),
            ("Path (optional)", self.batch_path_column, "path"),
        ]

        self.batch_mapping_combos = {}

        for row_num, (label_text, variable, key) in enumerate(mapping_fields):
            ttk.Label(
                mapping_frame,
                text=label_text + ":",
            ).grid(
                row=row_num,
                column=0,
                sticky="w",
                padx=(0, 8),
                pady=3,
            )

            combo = ttk.Combobox(
                mapping_frame,
                textvariable=variable,
                state="readonly",
            )
            combo.grid(
                row=row_num,
                column=1,
                sticky="ew",
                pady=3,
            )
            self.batch_mapping_combos[key] = combo

        mapping_frame.columnconfigure(1, weight=1)

        ttk.Button(
            mapping_frame,
            text="Apply Column Mapping",
            command=self.batch_apply_mapping,
        ).grid(
            row=len(mapping_fields),
            column=0,
            columnspan=2,
            sticky="e",
            pady=(8, 0),
        )


    def build_length_tab(self):
        main = self.length_tab

        controls = ttk.Frame(main)
        controls.pack(fill="x", pady=(0, 8))

        ttk.Button(
            controls,
            text="Add Excel Files",
            command=self.length_add_files,
        ).pack(side="left")

        ttk.Button(
            controls,
            text="Remove Selected",
            command=self.length_remove_selected,
        ).pack(side="left", padx=(8, 0))

        ttk.Button(
            controls,
            text="Clear List",
            command=self.length_clear,
        ).pack(side="left", padx=(8, 0))

        ttk.Label(
            controls,
            text="Show:",
        ).pack(side="left", padx=(20, 6))

        self.length_filter_combo = ttk.Combobox(
            controls,
            textvariable=self.length_filter,
            values=(
                "All",
            ),
            state="readonly",
            width=18,
        )
        self.length_filter_combo.pack(side="left")
        self.length_filter_combo.bind(
            "<<ComboboxSelected>>",
            self.length_filter_changed,
        )

        ttk.Button(
            controls,
            text="Calculate Lengths",
            command=self.length_calculate,
        ).pack(side="right")

        ttk.Button(
            controls,
            text="Export Cable List",
            command=self.length_export,
        ).pack(side="right", padx=(0, 8))

        list_frame = ttk.LabelFrame(
            main,
            text="Files and Worksheets",
            padding=8,
        )
        list_frame.pack(fill="both", expand=True, pady=(0, 8))

        columns = (
            "include",
            "file",
            "sheet",
            "building",
            "a",
            "z",
            "type",
            "path",
            "count",
        )

        self.length_tree = ttk.Treeview(
            list_frame,
            columns=columns,
            show="headings",
            selectmode="extended",
        )

        headings = {
            "include": "USE",
            "file": "File",
            "sheet": "Worksheet",
            "building": "Building",
            "a": "A Pos",
            "z": "Z Pos",
            "type": "Cable Type",
            "path": "Path",
            "count": "Cables",
        }

        for key in columns:
            self.length_tree.heading(key, text=headings[key])
            self.length_tree.column(
                key,
                width=90,
                anchor="center",
                stretch=True,
            )

        self.length_tree.column("include", width=90, stretch=False)
        self.length_tree.column("file", anchor="w")
        self.length_tree.column("sheet", anchor="w")

        length_style = ttk.Style()
        length_style.configure(
            "Length.Treeview",
            font=("TkDefaultFont", 10, "bold"),
            rowheight=26,
        )
        self.length_tree.configure(style="Length.Treeview")

        length_v = ttk.Scrollbar(
            list_frame,
            orient="vertical",
            command=self.length_tree.yview,
        )
        length_h = ttk.Scrollbar(
            list_frame,
            orient="horizontal",
            command=self.length_tree.xview,
        )
        self.length_tree.configure(
            yscrollcommand=length_v.set,
            xscrollcommand=length_h.set,
        )

        self.length_tree.grid(row=0, column=0, sticky="nsew")
        length_v.grid(row=0, column=1, sticky="ns")
        length_h.grid(row=1, column=0, sticky="ew")

        list_frame.rowconfigure(0, weight=1)
        list_frame.columnconfigure(0, weight=1)

        self.length_tree.bind(
            "<Button-1>",
            self.length_tree_click,
            add="+",
        )
        self.length_tree.bind(
            "<<TreeviewSelect>>",
            self.length_load_selected,
        )

        editor = ttk.Frame(main)
        editor.pack(fill="x", pady=(0, 8))

        settings_frame = ttk.LabelFrame(
            editor,
            text="Selected Worksheet Settings",
            padding=10,
        )
        settings_frame.pack(
            side="left",
            fill="both",
            expand=True,
            padx=(0, 5),
        )

        ttk.Label(
            settings_frame,
            text="Building:",
        ).grid(
            row=0,
            column=0,
            sticky="w",
            padx=(0, 8),
            pady=3,
        )

        self.length_building_combo = ttk.Combobox(
            settings_frame,
            textvariable=self.length_building,
            values=tuple(self.length_configs.keys()),
            state="readonly",
        )
        self.length_building_combo.grid(
            row=0,
            column=1,
            sticky="ew",
            pady=3,
        )

        settings_frame.columnconfigure(1, weight=1)

        ttk.Button(
            settings_frame,
            text="Apply to Selected",
            command=self.length_apply_selected,
        ).grid(
            row=1,
            column=0,
            columnspan=2,
            sticky="e",
            pady=(8, 0),
        )

        mapping_frame = ttk.LabelFrame(
            editor,
            text="Selected Worksheet Column Mapping",
            padding=10,
        )
        mapping_frame.pack(
            side="left",
            fill="both",
            expand=True,
            padx=(5, 0),
        )

        mapping_fields = [
            ("A Position", self.length_a_column, "a_position"),
            ("Z Position", self.length_z_column, "z_position"),
            ("Cable Type", self.length_cable_type_column, "cable_type"),
            ("Path (blank = A)", self.length_path_column, "path"),
        ]

        self.length_mapping_combos = {}

        for row_num, (label_text, variable, key) in enumerate(
            mapping_fields
        ):
            ttk.Label(
                mapping_frame,
                text=label_text + ":",
            ).grid(
                row=row_num,
                column=0,
                sticky="w",
                padx=(0, 8),
                pady=3,
            )

            combo = ttk.Combobox(
                mapping_frame,
                textvariable=variable,
                state="readonly",
            )
            combo.grid(
                row=row_num,
                column=1,
                sticky="ew",
                pady=3,
            )
            self.length_mapping_combos[key] = combo

        mapping_frame.columnconfigure(1, weight=1)

        ttk.Button(
            mapping_frame,
            text="Apply Column Mapping",
            command=self.length_apply_selected,
        ).grid(
            row=len(mapping_fields),
            column=0,
            columnspan=2,
            sticky="e",
            pady=(8, 0),
        )

        config_controls = ttk.Frame(editor)
        config_controls.pack(
            side="left",
            fill="y",
            padx=(10, 0),
        )

        self.length_edit_buildings_button = ttk.Button(
            config_controls,
            text="⚙",
            width=3,
            command=self.length_toggle_building_editor,
        )
        self.length_edit_buildings_button.pack()

        self.length_config_frame = ttk.LabelFrame(
            main,
            text="Length Configuration",
            padding=10,
        )
        config_frame = self.length_config_frame

        ttk.Label(
            config_frame,
            text="Edit Building:",
        ).grid(
            row=0,
            column=0,
            sticky="w",
            padx=(0, 8),
            pady=3,
        )

        self.length_config_building_combo = ttk.Combobox(
            config_frame,
            textvariable=self.length_config_building,
            values=tuple(self.length_configs.keys()),
            state="readonly",
        )
        config_building_combo = self.length_config_building_combo
        config_building_combo.grid(
            row=0,
            column=1,
            sticky="ew",
            pady=3,
        )
        config_building_combo.bind(
            "<<ComboboxSelected>>",
            self.length_load_config,
        )

        config_fields = [
            ("A Floor Location", self.length_a_floor),
            ("B Floor Location", self.length_b_floor),
            ("C Floor Location", self.length_c_floor),
            ("D Floor Location", self.length_d_floor),
            ("Distance to Wire Tray (m)", self.length_wire_tray),
            ("Distance Between Paired Rows (m)", self.length_paired_rows),
            ("Distance Between Aisles (m)", self.length_aisles),
        ]

        for row_num, (label_text, variable) in enumerate(
            config_fields,
            start=1,
        ):
            ttk.Label(
                config_frame,
                text=label_text + ":",
            ).grid(
                row=row_num,
                column=0,
                sticky="w",
                padx=(0, 8),
                pady=3,
            )

            ttk.Entry(
                config_frame,
                textvariable=variable,
            ).grid(
                row=row_num,
                column=1,
                sticky="ew",
                pady=3,
            )

        config_frame.columnconfigure(1, weight=1)

        config_buttons = ttk.Frame(config_frame)
        config_buttons.grid(
            row=8,
            column=0,
            columnspan=2,
            sticky="e",
            pady=(8, 0),
        )

        ttk.Button(
            config_buttons,
            text="Add Building",
            command=self.length_add_building,
        ).pack(side="left")

        ttk.Button(
            config_buttons,
            text="Apply Building Config",
            command=self.length_apply_config,
        ).pack(side="left", padx=(8, 0))

        preview = ttk.Notebook(main)
        self.length_preview_notebook = preview
        preview.pack(fill="both", expand=True)

        output_frame = ttk.Frame(preview)
        raw_frame = ttk.Frame(preview)

        preview.add(output_frame, text="Cable List")
        preview.add(raw_frame, text="Raw Cable Lengths")

        self.length_output_tree = ttk.Treeview(
            output_frame,
            show="headings",
        )
        self.length_raw_tree = ttk.Treeview(
            raw_frame,
            show="headings",
        )

        for frame, tree in (
            (output_frame, self.length_output_tree),
            (raw_frame, self.length_raw_tree),
        ):
            vbar = ttk.Scrollbar(
                frame,
                orient="vertical",
                command=tree.yview,
            )
            hbar = ttk.Scrollbar(
                frame,
                orient="horizontal",
                command=tree.xview,
            )
            tree.configure(
                yscrollcommand=vbar.set,
                xscrollcommand=hbar.set,
            )

            tree.grid(row=0, column=0, sticky="nsew")
            vbar.grid(row=0, column=1, sticky="ns")
            hbar.grid(row=1, column=0, sticky="ew")

            frame.rowconfigure(0, weight=1)
            frame.columnconfigure(0, weight=1)

        self.length_load_config()

    def length_toggle_building_editor(self):
        if self.length_building_editor_visible:
            self.length_config_frame.pack_forget()
            self.length_building_editor_visible = False
            self.length_edit_buildings_button.configure(
                text="⚙"
            )
        else:
            self.length_config_frame.pack(
                fill="x",
                pady=(0, 8),
                before=self.length_preview_notebook,
            )
            self.length_building_editor_visible = True
            self.length_edit_buildings_button.configure(
                text="⚙"
            )

    def length_refresh_building_choices(self):
        values = tuple(self.length_configs.keys())
        self.length_building_combo["values"] = values
        self.length_config_building_combo["values"] = values

    def length_add_building(self):
        name = simpledialog.askstring(
            APP_NAME,
            "New building name:",
            parent=self.root,
        )

        if not name:
            return

        name = name.strip()
        if not name:
            return

        if name in self.length_configs:
            messagebox.showwarning(
                APP_NAME,
                f'Building "{name}" already exists.',
            )
            return

        source_name = self.length_config_building.get() or "Default"
        source_config = self.length_configs.get(
            source_name,
            self.length_configs["Default"],
        )

        self.length_configs[name] = dict(source_config)
        self.length_refresh_building_choices()

        self.length_config_building.set(name)
        self.length_building.set(name)
        self.length_load_config()

        if not self.length_building_editor_visible:
            self.length_toggle_building_editor()

        self.status_text.set(f'Added Building {name}.')

    def length_collect_loaded_cable_types(self):
        cable_types = []

        for item in self.length_items:
            if not item.get("include", True):
                continue

            cable_types.extend(
                item.get("cached_cable_types", [])
            )

        return cable_types


    def length_refresh_filter_choices(self):
        cable_types = self.length_collect_loaded_cable_types()

        fiber_types = set()
        copper_types = set()
        has_other = False

        for cable_type in cable_types:
            category = cable_length_category(cable_type)
            subtype = cable_length_subtype(cable_type)

            if category == "Fiber":
                fiber_types.add(subtype)

            elif category == "Copper":
                copper_types.add(subtype)

            else:
                has_other = True

        def fiber_sort_key(value):
            match = re.match(r"(\d+)F$", value, re.I)
            if match:
                return (0, int(match.group(1)))
            return (1, value.upper())

        preferred_copper_order = {
            "Orange Copper": 0,
            "Green Copper": 1,
            "Copper": 2,
        }

        copper_list = sorted(
            copper_types,
            key=lambda value: (
                preferred_copper_order.get(value, 50),
                value.upper(),
            ),
        )

        values = [
            "All",
        ]

        if fiber_types:
            values.append("Fiber")

        if copper_types:
            values.append("Copper")

        values.extend(copper_list)
        values.extend(sorted(fiber_types, key=fiber_sort_key))

        if has_other:
            values.append("Other")

        # Remove duplicates while preserving order.
        unique_values = []
        seen = set()

        for value in values:
            if value not in seen:
                seen.add(value)
                unique_values.append(value)

        self.length_filter_combo["values"] = tuple(unique_values)

        current = self.length_filter.get()

        if current not in unique_values:
            self.length_filter.set("All")

    def length_add_files(self):
        selected = filedialog.askopenfilenames(
            title="Select Excel Cable Files",
            filetypes=[
                ("Excel Workbooks", "*.xlsx *.xlsm"),
                ("Excel Workbook", "*.xlsx"),
                ("Excel Macro-Enabled Workbook", "*.xlsm"),
                ("All Files", "*.*"),
            ],
        )

        if not selected:
            return

        self.start_loading("Loading...")

        try:
            for file_path in selected:
                if file_path in self.length_file_paths:
                    continue

                overview = self.get_cached_workbook_overview(
                    file_path
                )

                for sheet_name in overview.get(
                    "sheet_names",
                    [],
                ):
                    info = overview.get(
                        "sheets",
                        {},
                    ).get(
                        sheet_name,
                        {},
                    )

                    header_row = info.get("header_row")
                    columns = dict(
                        info.get(
                            "auto_columns",
                            {},
                        )
                    )

                    if not columns:
                        continue

                    try:
                        labels = self.cached_extract_labels(
                            file_path,
                            sheet_name,
                            header_row_override=header_row,
                            columns_override=columns,
                        )
                    except Exception:
                        labels = []

                    cached_cable_types = [
                        clean_value(
                            label.get(
                                "cable_type",
                                "",
                            )
                        )
                        for label in labels
                        if clean_value(
                            label.get(
                                "cable_type",
                                "",
                            )
                        )
                    ]

                    cached_column_choices = list(
                        info.get(
                            "choices",
                            [],
                        )
                    )
                    cached_choice_by_col = dict(
                        info.get(
                            "choice_by_col",
                            {},
                        )
                    )

                    building = infer_building_from_text(
                        os.path.basename(file_path),
                        sheet_name,
                    )

                    length_item = {
                        "include": True,
                        "file_path": file_path,
                        "file_name": os.path.basename(
                            file_path
                        ),
                        "sheet": sheet_name,
                        "header_row": header_row,
                        "columns": columns,
                        "building": building,
                        "count": len(labels),
                        "project_number": "",
                        "identifier": "",
                        "other": "",
                        "cached_cable_types": cached_cable_types,
                        "cached_column_choices": cached_column_choices,
                        "cached_choice_by_col": cached_choice_by_col,
                    }

                    batch_metadata = (
                        self.find_batch_metadata_for_sheet(
                            length_item
                        )
                    )

                    if batch_metadata is not length_item:
                        length_item["project_number"] = (
                            batch_metadata.get(
                                "project_number",
                                "",
                            )
                        )
                        length_item["identifier"] = (
                            batch_metadata.get(
                                "identifier",
                                "",
                            )
                        )
                        length_item["other"] = (
                            batch_metadata.get(
                                "other",
                                "",
                            )
                        )

                    self.length_items.append(length_item)

                self.length_file_paths.append(file_path)

            self.length_rebuild_tree()
            self.length_refresh_filter_choices()

            self.status_text.set(
                f"Cable Lengths: "
                f"{len(self.length_items)} worksheet(s) loaded."
            )

        except Exception as exc:
            messagebox.showerror(
                APP_NAME,
                f"Could not load cable-length files.\n\n{exc}",
            )
        finally:
            self.stop_loading()


    def length_remove_selected(self):
        selected = list(self.length_tree.selection())
        if not selected:
            return

        for index in sorted(
            (int(item_id) for item_id in selected),
            reverse=True,
        ):
            if 0 <= index < len(self.length_items):
                del self.length_items[index]

        self.length_rebuild_tree()
        self.length_refresh_filter_choices()

    def length_clear(self):
        self.length_items = []
        self.length_file_paths = []
        self.length_selected_index = None
        self.length_raw_records = []
        self.length_summary_rows = []

        self.clear_tree(self.length_tree)
        self.clear_tree(self.length_output_tree)
        self.clear_tree(self.length_raw_tree)

        self.length_filter.set("All")
        self.length_filter_combo["values"] = (
            "All",
        )

        self.status_text.set("Cable Lengths list cleared.")

    def length_rebuild_tree(self, select_index=None):
        self.clear_tree(self.length_tree)
        self.length_file_paths = []

        for index, item in enumerate(self.length_items):
            if item["file_path"] not in self.length_file_paths:
                self.length_file_paths.append(item["file_path"])

            columns = item["columns"]

            self.length_tree.insert(
                "",
                "end",
                iid=str(index),
                values=[
                    "☑ USE" if item["include"] else "☐ SKIP",
                    item["file_name"],
                    item["sheet"],
                    item["building"],
                    get_column_letter(columns["a_position"]),
                    get_column_letter(columns["z_position"]),
                    get_column_letter(columns["cable_type"]),
                    (
                        get_column_letter(columns["path"])
                        if columns.get("path")
                        else "-"
                    ),
                    item["count"],
                ],
            )

        self.autosize_tree_columns(
            self.length_tree,
            min_width=55,
            max_width=240,
        )
        self.length_tree.column("include", width=90, stretch=False)

        if (
            select_index is not None
            and 0 <= select_index < len(self.length_items)
        ):
            iid = str(select_index)
            self.length_tree.selection_set(iid)
            self.length_tree.focus(iid)
            self.length_tree.see(iid)

    def length_tree_click(self, event):
        if self.length_tree.identify_region(event.x, event.y) != "cell":
            return

        row_id = self.length_tree.identify_row(event.y)
        column_id = self.length_tree.identify_column(event.x)

        if not row_id or column_id != "#1":
            return

        index = int(row_id)
        item = self.length_items[index]
        item["include"] = not item["include"]

        self.length_rebuild_tree(select_index=index)
        self.length_refresh_filter_choices()
        return "break"

    def length_load_selected(self, event=None):
        selection = self.length_tree.selection()
        if not selection:
            self.length_selected_index = None
            return

        index = int(selection[0])
        if index < 0 or index >= len(self.length_items):
            return

        self.length_selected_index = index
        item = self.length_items[index]

        self.length_building.set(item["building"])
        self.length_setup_mapping_choices(item)

    def length_setup_mapping_choices(self, item):
        try:
            choices = list(
                item.get("cached_column_choices", [])
            )
            choice_by_col = dict(
                item.get("cached_choice_by_col", {})
            )

            if not choices:
                info = self.get_cached_sheet_info(
                    item["file_path"],
                    item["sheet"],
                )
                choices = list(info.get("choices", []))
                choice_by_col = dict(
                    info.get("choice_by_col", {})
                )

                item["cached_column_choices"] = list(choices)
                item["cached_choice_by_col"] = dict(
                    choice_by_col
                )

            variables = {
                "a_position": self.length_a_column,
                "z_position": self.length_z_column,
                "cable_type": self.length_cable_type_column,
                "path": self.length_path_column,
            }

            for key, combo in self.length_mapping_combos.items():
                combo["values"] = list(choices)

                col_num = item["columns"].get(key)

                if col_num:
                    variables[key].set(
                        choice_by_col.get(
                            col_num,
                            "",
                        )
                    )
                else:
                    variables[key].set("")

        except Exception as exc:
            messagebox.showerror(
                APP_NAME,
                f"Could not load worksheet columns.\n\n{exc}",
            )


    def length_get_selected_columns(self):
        columns = {
            "a_position": self.choice_to_column(
                self.length_a_column.get()
            ),
            "z_position": self.choice_to_column(
                self.length_z_column.get()
            ),
            "cable_type": self.choice_to_column(
                self.length_cable_type_column.get()
            ),
        }

        path_col = self.choice_to_column(
            self.length_path_column.get()
        )

        if path_col:
            columns["path"] = path_col
        else:
            columns.pop("path", None)

        if not all(
            columns.get(key)
            for key in (
                "a_position",
                "z_position",
                "cable_type",
            )
        ):
            raise RuntimeError(
                "Select A Position, Z Position, and Cable Type columns."
            )

        return columns

    def length_apply_selected(self):
        if self.length_selected_index is None:
            messagebox.showwarning(
                APP_NAME,
                "Select a worksheet first.",
            )
            return

        try:
            columns = self.length_get_selected_columns()
        except Exception as exc:
            messagebox.showwarning(APP_NAME, str(exc))
            return

        item = self.length_items[self.length_selected_index]
        item["columns"] = columns
        item["building"] = self.length_building.get()

        try:
            labels = self.cached_extract_labels(
                item["file_path"],
                item["sheet"],
                header_row_override=item["header_row"],
                columns_override=columns,
            )
            item["count"] = len(labels)
            item["cached_cable_types"] = [
                clean_value(label.get("cable_type", ""))
                for label in labels
                if clean_value(label.get("cable_type", ""))
            ]
        except Exception as exc:
            messagebox.showerror(
                APP_NAME,
                f"Could not apply worksheet settings.\n\n{exc}",
            )
            return

        self.length_rebuild_tree(
            select_index=self.length_selected_index
        )
        self.length_refresh_filter_choices()

    def length_load_config(self, event=None):
        building = self.length_config_building.get() or "Default"
        config = self.length_configs[building]

        self.length_a_floor.set(str(config["a_floor"]))
        self.length_b_floor.set(str(config["b_floor"]))
        self.length_c_floor.set(str(config["c_floor"]))
        self.length_d_floor.set(str(config["d_floor"]))
        self.length_wire_tray.set(str(config["distance_to_wire_tray"]))
        self.length_paired_rows.set(str(config["row_pair_spacing"]))
        self.length_aisles.set(str(config["between_pair_spacing"]))

    def length_apply_config(self):
        building = self.length_config_building.get() or "Default"

        try:
            config = self.length_configs[building]
            config["a_floor"] = int(self.length_a_floor.get())
            config["b_floor"] = int(self.length_b_floor.get())
            config["c_floor"] = int(self.length_c_floor.get())
            config["d_floor"] = int(self.length_d_floor.get())
            config["distance_to_wire_tray"] = float(
                self.length_wire_tray.get()
            )
            config["row_pair_spacing"] = float(
                self.length_paired_rows.get()
            )
            config["between_pair_spacing"] = float(
                self.length_aisles.get()
            )

            if building == "100":
                config["romp_mode"] = "building100TwoRomp"
            else:
                config["romp_mode"] = "none"

        except ValueError:
            messagebox.showwarning(
                APP_NAME,
                "Building configuration values must be numeric.",
            )
            return

        self.status_text.set(
            f"Cable length configuration saved for Building {building}."
        )

    def length_make_records(self):
        records = []

        for item in self.length_items:
            if not item["include"]:
                continue

            config = self.length_configs[item["building"]]

            labels = self.cached_extract_labels(
                item["file_path"],
                item["sheet"],
                header_row_override=item["header_row"],
                columns_override=item["columns"],
            )

            for label in labels:
                a_location = parse_bulk_location(label["a_position"])
                z_location = parse_bulk_location(label["z_position"])

                record = {
                    "file_name": item["file_name"],
                    "sheet": item["sheet"],
                    "building": item["building"],
                    "a_position": label["a_position"],
                    "z_position": label["z_position"],
                    "path": extract_path_letter(label["path"]),
                    "cable_type": label["cable_type"],
                    "raw_length_m": None,
                    "rounded_length_m": None,
                    "status": "OK",
                }

                if not a_location or not z_location:
                    record["status"] = "Invalid A/Z location"
                    records.append(record)
                    continue

                side_a = infer_location_side(label["a_position"])
                side_z = infer_location_side(label["z_position"])
                room_a = infer_romp(label["a_position"])
                room_z = infer_romp(label["z_position"])

                path = record["path"]

                # Path is required as a worksheet column. If an individual
                # cable's Path cell is blank, calculate that cable as Path A.
                if not path:
                    path = "A"
                    record["path"] = path

                raw_distance = calculate_bulk_distance(
                    a_location,
                    z_location,
                    path,
                    config,
                    side_a,
                    side_z,
                    room_a,
                    room_z,
                )

                if raw_distance == "INVALID_FLOOR":
                    record["status"] = "Floor outside configured range"
                    records.append(record)
                    continue

                if raw_distance is None:
                    record["status"] = "Could not calculate path"
                    records.append(record)
                    continue

                record["raw_length_m"] = float(raw_distance)

                if "COPPER" in record["cable_type"].upper():
                    record["rounded_length_m"] = meters_to_whole_feet(
                        raw_distance
                    )
                else:
                    record["rounded_length_m"] = round_up_to_nearest_five(
                        raw_distance
                    )

                records.append(record)

        return records

    def length_filtered_records(self):
        filter_name = self.length_filter.get() or "All"

        return [
            record
            for record in self.length_raw_records
            if cable_matches_length_filter(
                record.get("cable_type", ""),
                filter_name,
            )
        ]

    def length_filter_changed(self, event=None):
        if not self.length_raw_records:
            return

        filtered = self.length_filtered_records()
        self.length_summary_rows = summarize_cable_lengths(filtered)
        self.length_refresh_preview()

        self.status_text.set(
            f'Cable Lengths filter: {self.length_filter.get()}'
        )

    def length_calculate(self):
        if not any(item["include"] for item in self.length_items):
            messagebox.showwarning(
                APP_NAME,
                "Add and include at least one worksheet first.",
            )
            return

        self.start_loading("Loading...")

        try:
            self.length_raw_records = self.length_make_records()
            self.length_summary_rows = summarize_cable_lengths(
                self.length_filtered_records()
            )

            self.length_refresh_preview()

            valid = sum(
                1
                for record in self.length_raw_records
                if record["status"] == "OK"
            )
            errors = len(self.length_raw_records) - valid

            self.status_text.set(
                f"Cable Lengths: {valid} calculated"
                + (f", {errors} issue(s)" if errors else "")
            )

        except Exception as exc:
            messagebox.showerror(
                APP_NAME,
                f"Could not calculate cable lengths.\n\n{exc}",
            )
        finally:
            self.stop_loading()

    def length_refresh_preview(self):
        self.clear_tree(self.length_output_tree)
        self.clear_tree(self.length_raw_tree)

        summary_headers = [
            "Cable Type",
            "Length",
            "Quantity",
            "Total Length",
        ]
        summary_ids = [
            f"summary_{index}"
            for index in range(len(summary_headers))
        ]

        self.length_output_tree["columns"] = summary_ids

        for column_id, header in zip(summary_ids, summary_headers):
            self.length_output_tree.heading(column_id, text=header)
            self.length_output_tree.column(
                column_id,
                anchor="w",
                width=100,
            )

        for row in self.length_summary_rows:
            if row.get("separator"):
                self.length_output_tree.insert(
                    "",
                    "end",
                    values=["", "", "", ""],
                )
                continue

            self.length_output_tree.insert(
                "",
                "end",
                values=[
                    row["cable_type"],
                    row["length"],
                    row["quantity"],
                    row["total_length"],
                ],
            )

        raw_headers = [
            "File",
            "Worksheet",
            "Building",
            "A Position",
            "Z Position",
            "Path",
            "Cable Type",
            "Raw Length",
            "Rounded Length",
            "Status",
        ]
        raw_ids = [
            f"raw_{index}"
            for index in range(len(raw_headers))
        ]

        self.length_raw_tree["columns"] = raw_ids

        for column_id, header in zip(raw_ids, raw_headers):
            self.length_raw_tree.heading(column_id, text=header)
            self.length_raw_tree.column(
                column_id,
                anchor="w",
                width=100,
            )

        filtered_records = self.length_filtered_records()

        # Keep copper below all non-copper in the raw preview as well.
        filtered_records = sorted(
            filtered_records,
            key=lambda record: (
                1 if "COPPER" in record["cable_type"].upper() else 0,
                record["cable_type"].upper(),
                record["file_name"].upper(),
                record["sheet"].upper(),
            ),
        )

        inserted_copper_separator = False

        for record in filtered_records:
            is_copper = "COPPER" in record["cable_type"].upper()

            if (
                is_copper
                and not inserted_copper_separator
                and any(
                    "COPPER" not in other["cable_type"].upper()
                    for other in filtered_records
                )
            ):
                self.length_raw_tree.insert(
                    "",
                    "end",
                    values=["", "", "", "", "", "", "", "", "", ""],
                )
                inserted_copper_separator = True

            if is_copper and isinstance(record.get("raw_length_m"), (int, float)):
                raw_display = f'{meters_to_whole_feet(record["raw_length_m"])} ft'
                rounded_display = raw_display
            else:
                raw_display = (
                    f'{round(record["raw_length_m"], 2)} m'
                    if isinstance(record.get("raw_length_m"), (int, float))
                    else ""
                )
                rounded_display = (
                    f'{record["rounded_length_m"]} m'
                    if record.get("rounded_length_m") is not None
                    else ""
                )

            self.length_raw_tree.insert(
                "",
                "end",
                values=[
                    record["file_name"],
                    record["sheet"],
                    record["building"],
                    record["a_position"],
                    record["z_position"],
                    record["path"],
                    record["cable_type"],
                    raw_display,
                    rounded_display,
                    record["status"],
                ],
            )

        self.autosize_tree_columns(
            self.length_output_tree,
            min_width=70,
            max_width=320,
        )
        self.autosize_tree_columns(
            self.length_raw_tree,
            min_width=70,
            max_width=320,
        )


    def length_export(self):
        if not self.length_raw_records:
            self.length_calculate()

            if not self.length_raw_records:
                return

        output_path = filedialog.asksaveasfilename(
            title="Save Cable Length List",
            initialfile=self.first_length_output_filename(),
            defaultextension=".xlsx",
            filetypes=[("Excel Workbook", "*.xlsx")],
        )

        if not output_path:
            return

        self.start_loading("Loading...")

        try:
            filtered_records = self.length_filtered_records()
            filtered_summary = summarize_cable_lengths(filtered_records)

            write_cable_length_output(
                output_path,
                filtered_summary,
                filtered_records,
            )

            messagebox.showinfo(
                APP_NAME,
                "Cable length list exported successfully.",
            )
        except Exception as exc:
            messagebox.showerror(
                APP_NAME,
                f"Could not export cable length list.\n\n{exc}",
            )
        finally:
            self.stop_loading()

    def clear_tree(self, tree):
        for item in tree.get_children():
            tree.delete(item)

    def autosize_tree_columns(self, tree, min_width=55, max_width=320):
        font = tkfont.nametofont("TkDefaultFont")

        for column in tree["columns"]:
            heading = tree.heading(column, "text")
            width = font.measure(str(heading)) + 24

            for item_id in tree.get_children():
                value = tree.set(item_id, column)
                # Treeview cells are single-line. Size to the longest line.
                lines = str(value).splitlines() or [""]
                for line in lines:
                    width = max(width, font.measure(line) + 24)

            tree.column(
                column,
                width=max(min_width, min(width, max_width)),
                stretch=True,
            )


    def on_output_type_changed(self, event=None):
        self.start_loading("Changing output type...")
        try:
            self.load_output_preview()
            self.pulse_loading()
        finally:
            self.stop_loading()

    def generate_selected_output(self):
        output_type = self.output_type.get()

        if output_type == "Printable Labels":
            self.generate_printable_labels()
        elif output_type == "Easy-Mark":
            self.generate_easymark_file()
        else:
            self.generate_labels()

    def get_single_preview_data(self):
        source = self.file_path.get().strip()
        sheet_name = self.sheet_name.get().strip()

        if not source or not sheet_name or not self.current_header_row:
            return []

        columns = self.get_selected_columns()

        return self.cached_extract_labels(
            source,
            sheet_name,
            header_row_override=self.current_header_row,
            columns_override=columns,
            progress_callback=self.pulse_loading,
        )

    def load_output_preview(self):
        self.clear_tree(self.output_preview)

        try:
            labels = self.get_single_preview_data()
            start_number = int(
                self.starting_cable_number.get().strip()
            )
        except Exception:
            self.output_preview["columns"] = ()
            return

        output_type = self.output_type.get()
        project = self.project_number.get().strip()
        identifier = self.identifier.get().strip()
        other = self.other.get().strip()

        if output_type == "Easy-Mark":
            headers = [
                "Cable Number - Path",
                "A Position",
                "Z Position",
                "Cable Type",
                "Project Number - Identifier",
                "Other",
            ]
            column_ids = [
                f"c{i}"
                for i in range(len(headers))
            ]
            self.output_preview["columns"] = column_ids

            for col_id, header in zip(column_ids, headers):
                self.output_preview.heading(
                    col_id,
                    text=header,
                )
                self.output_preview.column(
                    col_id,
                    anchor="w",
                    width=100,
                )

            for index, label in enumerate(labels[:200]):
                cable_number = start_number + index

                number_path = (
                    f"{cable_number} - {label['path']}"
                    if label["path"]
                    else str(cable_number)
                )

                project_identifier = project
                if project and identifier:
                    project_identifier = (
                        f"{project} - {identifier}"
                    )
                elif identifier:
                    project_identifier = identifier

                self.output_preview.insert(
                    "",
                    "end",
                    values=[
                        number_path,
                        label["a_position"],
                        label["z_position"],
                        label["cable_type"],
                        project_identifier,
                        other,
                    ],
                )

        elif output_type == "Printable Labels":
            headers = [
                "Path",
                "Project Number",
                "Identifier",
                "Other",
                "A Position",
                "Cable Number",
                "Cable Type",
                "Z Position",
            ]
            column_ids = [
                f"c{i}"
                for i in range(len(headers))
            ]
            self.output_preview["columns"] = column_ids

            for col_id, header in zip(column_ids, headers):
                self.output_preview.heading(
                    col_id,
                    text=header,
                )
                self.output_preview.column(
                    col_id,
                    anchor="w",
                    width=100,
                )

            for index, label in enumerate(labels[:200]):
                self.output_preview.insert(
                    "",
                    "end",
                    values=[
                        label["path"],
                        project,
                        identifier,
                        other,
                        label["a_position"],
                        start_number + index,
                        label["cable_type"],
                        label["z_position"],
                    ],
                )

        else:
            headers = [
                "A Position",
                "Project Number",
                "Cable Number",
                "Identifier",
                "Other",
                "Cable Type",
                "Z Position",
                "Path",
            ]
            column_ids = [
                f"c{i}"
                for i in range(len(headers))
            ]
            self.output_preview["columns"] = column_ids

            for col_id, header in zip(column_ids, headers):
                self.output_preview.heading(
                    col_id,
                    text=header,
                )
                self.output_preview.column(
                    col_id,
                    anchor="w",
                    width=100,
                )

            for index, label in enumerate(labels[:200]):
                self.output_preview.insert(
                    "",
                    "end",
                    values=[
                        label["a_position"],
                        project,
                        start_number + index,
                        identifier,
                        other,
                        label["cable_type"],
                        label["z_position"],
                        label["path"],
                    ],
                )

        self.autosize_tree_columns(
            self.output_preview,
            min_width=70,
            max_width=300,
        )


    def refresh_previews(self):
        self.start_loading("Refreshing preview...")
        try:
            self.load_output_preview()
        except Exception as exc:
            messagebox.showerror(
                APP_NAME,
                f"Could not refresh preview.\n\n{exc}",
            )
        finally:
            self.stop_loading()


    def batch_add_files(self):
        selected = filedialog.askopenfilenames(
            title="Select Excel Cable Files",
            filetypes=[
                ("Excel Workbooks", "*.xlsx *.xlsm"),
                ("Excel Workbook", "*.xlsx"),
                ("Excel Macro-Enabled Workbook", "*.xlsm"),
                ("All Files", "*.*"),
            ],
        )

        if not selected:
            return

        self.start_loading("Loading files...")

        try:
            for file_path in selected:
                if file_path in self.batch_file_paths:
                    continue

                try:
                    overview = self.get_cached_workbook_overview(
                        file_path
                    )

                    for sheet_name in overview.get(
                        "sheet_names",
                        [],
                    ):
                        info = overview.get(
                            "sheets",
                            {},
                        ).get(
                            sheet_name,
                            {},
                        )

                        header_row = info.get("header_row")
                        columns = dict(
                            info.get(
                                "auto_columns",
                                {},
                            )
                        )

                        if not columns:
                            continue

                        try:
                            labels = self.cached_extract_labels(
                                file_path,
                                sheet_name,
                                header_row_override=header_row,
                                columns_override=columns,
                                progress_callback=self.pulse_loading,
                            )
                        except Exception:
                            labels = []

                        self.batch_items.append(
                            {
                                "include": True,
                                "file_path": file_path,
                                "file_name": os.path.basename(
                                    file_path
                                ),
                                "sheet": sheet_name,
                                "header_row": header_row,
                                "columns": columns,
                                "count": len(labels),
                                "project_number": "",
                                "cable_start": 1,
                                "identifier": "",
                                "other": "",
                            }
                        )

                    self.batch_file_paths.append(file_path)

                except Exception as exc:
                    messagebox.showerror(
                        APP_NAME,
                        f"Could not read:\n{file_path}\n\n{exc}",
                    )

            self.rebuild_batch_tree()

            self.status_text.set(
                f"Batch list contains "
                f"{len(self.batch_items)} worksheet(s)."
            )

        finally:
            self.stop_loading()


    def batch_load_selected_settings(self, event=None):
        selection = self.batch_tree.selection()

        if not selection:
            self.batch_selected_index = None
            return

        index = int(selection[0])

        if index < 0 or index >= len(self.batch_items):
            return

        self.batch_selected_index = index
        item = self.batch_items[index]

        self.batch_project_number.set(
            item.get("project_number", "")
        )
        self.batch_cable_start.set(
            str(item.get("cable_start", 1))
        )
        self.batch_identifier.set(
            item.get("identifier", "")
        )
        self.batch_other.set(
            item.get("other", "")
        )

        self.batch_setup_mapping_choices(item)

    def batch_setup_mapping_choices(self, item):
        try:
            info = self.get_cached_sheet_info(
                item["file_path"],
                item["sheet"],
            )

            choices = list(info.get("choices", []))
            choice_by_col = dict(
                info.get("choice_by_col", {})
            )

            variables = {
                "a_position": self.batch_a_column,
                "z_position": self.batch_z_column,
                "cable_type": self.batch_cable_type_column,
                "path": self.batch_path_column,
            }

            for key, combo in self.batch_mapping_combos.items():
                combo_values = list(choices)

                if key == "path":
                    combo_values = ["(None)"] + combo_values

                combo["values"] = combo_values

                col_num = item["columns"].get(key)

                if col_num:
                    variables[key].set(
                        choice_by_col.get(
                            col_num,
                            "",
                        )
                    )
                elif key == "path":
                    variables[key].set("(None)")
                else:
                    variables[key].set("")

        except Exception as exc:
            messagebox.showerror(
                APP_NAME,
                f"Could not load column mapping.\n\n{exc}",
            )


    def get_batch_selected_columns(self):
        columns = {
            "a_position": self.choice_to_column(
                self.batch_a_column.get()
            ),
            "z_position": self.choice_to_column(
                self.batch_z_column.get()
            ),
            "cable_type": self.choice_to_column(
                self.batch_cable_type_column.get()
            ),
        }

        path_col = self.choice_to_column(
            self.batch_path_column.get()
        )

        if path_col:
            columns["path"] = path_col

        if not all(
            columns.get(key)
            for key in (
                "a_position",
                "z_position",
                "cable_type",
            )
        ):
            raise RuntimeError(
                "Select columns for A Position, Z Position, "
                "and Cable Type."
            )

        return columns

    def get_batch_settings_values(self):
        try:
            cable_start = int(
                self.batch_cable_start.get().strip()
            )
            if cable_start < 0:
                raise ValueError
        except ValueError:
            raise RuntimeError(
                "Cable Start must be a whole number of 0 or greater."
            )

        return {
            "project_number": self.batch_project_number.get().strip(),
            "cable_start": cable_start,
            "identifier": self.batch_identifier.get().strip(),
            "other": self.batch_other.get().strip(),
        }

    def batch_apply_settings(self):
        if self.batch_selected_index is None:
            messagebox.showwarning(
                APP_NAME,
                "Select a worksheet in the list first.",
            )
            return

        try:
            values = self.get_batch_settings_values()
        except Exception as exc:
            messagebox.showwarning(
                APP_NAME,
                str(exc),
            )
            return

        item = self.batch_items[
            self.batch_selected_index
        ]
        item.update(values)

        self.rebuild_batch_tree(
            select_index=self.batch_selected_index
        )

        self.status_text.set(
            f'Settings saved for '
            f'{item["file_name"]} / {item["sheet"]}.'
        )

    def batch_apply_settings_all(self):
        if not self.batch_items:
            messagebox.showwarning(
                APP_NAME,
                "Add at least one worksheet first.",
            )
            return

        try:
            values = self.get_batch_settings_values()
        except Exception as exc:
            messagebox.showwarning(
                APP_NAME,
                str(exc),
            )
            return

        for item in self.batch_items:
            item.update(values)

        selected_index = self.batch_selected_index

        self.rebuild_batch_tree(
            select_index=selected_index
        )

        self.status_text.set(
            f"Settings applied to all "
            f"{len(self.batch_items)} worksheet(s)."
        )

    def batch_apply_mapping(self):
        if self.batch_selected_index is None:
            messagebox.showwarning(
                APP_NAME,
                "Select a worksheet in the list first.",
            )
            return

        try:
            columns = self.get_batch_selected_columns()
        except Exception as exc:
            messagebox.showwarning(
                APP_NAME,
                str(exc),
            )
            return

        item = self.batch_items[
            self.batch_selected_index
        ]

        self.start_loading("Applying mapping...")

        try:
            item["columns"] = columns

            labels = self.cached_extract_labels(
                item["file_path"],
                item["sheet"],
                header_row_override=item["header_row"],
                columns_override=columns,
                progress_callback=self.pulse_loading,
            )
            item["count"] = len(labels)

            self.rebuild_batch_tree(
                select_index=self.batch_selected_index
            )

            self.status_text.set(
                f'Column mapping saved for '
                f'{item["file_name"]} / {item["sheet"]}.'
            )

        except Exception as exc:
            messagebox.showerror(
                APP_NAME,
                f"Could not apply column mapping.\n\n{exc}",
            )
        finally:
            self.stop_loading()

    def batch_tree_click(self, event):
        region = self.batch_tree.identify_region(
            event.x,
            event.y,
        )

        if region != "cell":
            return

        row_id = self.batch_tree.identify_row(event.y)
        column_id = self.batch_tree.identify_column(event.x)

        if not row_id or column_id != "#1":
            return

        index = int(row_id)

        if index < 0 or index >= len(self.batch_items):
            return

        item = self.batch_items[index]
        item["include"] = not item["include"]

        self.rebuild_batch_tree(
            select_index=index
        )

        return "break"


    def on_batch_output_type_changed(self, event=None):
        self.start_loading("Changing output type...")
        try:
            self.pulse_loading()
            self.pulse_loading()
        finally:
            self.stop_loading()

    def generate_selected_batch_output(self):
        output_type = self.batch_output_type.get()

        if output_type == "Printable Labels":
            self.generate_batch_printable()
        elif output_type == "Easy-Mark":
            self.generate_batch_easymark()
        else:
            self.generate_batch_labels()



    def batch_remove_selected(self):
        selected = list(self.batch_tree.selection())
        if not selected:
            return

        remove_indexes = sorted((int(iid) for iid in selected), reverse=True)
        for index in remove_indexes:
            del self.batch_items[index]

        self.rebuild_batch_tree()

    def batch_clear(self):
        self.batch_items = []
        self.batch_file_paths = []
        self.batch_selected_index = None
        self.batch_project_number.set("")
        self.batch_cable_start.set("1")
        self.batch_identifier.set("")
        self.batch_other.set("")
        self.clear_tree(self.batch_tree)
        self.status_text.set("Batch list cleared.")

    def rebuild_batch_tree(self, select_index=None):
        self.clear_tree(self.batch_tree)
        self.batch_file_paths = []

        for index, item in enumerate(self.batch_items):
            if item["file_path"] not in self.batch_file_paths:
                self.batch_file_paths.append(
                    item["file_path"]
                )

            columns = item["columns"]

            self.batch_tree.insert(
                "",
                "end",
                iid=str(index),
                values=[
                    "☑  USE" if item["include"] else "☐  SKIP",
                    item["file_name"],
                    item["sheet"],
                    item.get("project_number", ""),
                    item.get("cable_start", 1),
                    item.get("identifier", ""),
                    item.get("other", ""),
                    get_column_letter(
                        columns["a_position"]
                    ),
                    get_column_letter(
                        columns["z_position"]
                    ),
                    get_column_letter(
                        columns["cable_type"]
                    ),
                    (
                        get_column_letter(
                            columns["path"]
                        )
                        if columns.get("path")
                        else "-"
                    ),
                    item["count"],
                ],
            )

        self.autosize_tree_columns(
            self.batch_tree,
            min_width=55,
            max_width=240,
        )

        self.batch_tree.column(
            "include",
            width=90,
            stretch=False,
        )

        if (
            select_index is not None
            and 0 <= select_index < len(
                self.batch_items
            )
        ):
            iid = str(select_index)
            self.batch_tree.selection_set(iid)
            self.batch_tree.focus(iid)
            self.batch_tree.see(iid)


    def collect_batch_sections(self):
        sections = []

        for item in self.batch_items:
            self.pulse_loading()

            if not item["include"]:
                continue

            labels = self.cached_extract_labels(
                item["file_path"],
                item["sheet"],
                header_row_override=item["header_row"],
                columns_override=item["columns"],
                progress_callback=self.pulse_loading,
            )

            if labels:
                sections.append(
                    {
                        "item": item,
                        "labels": labels,
                    }
                )

        if not sections:
            raise RuntimeError("No included cable labels were found.")

        return sections


    def generate_batch_labels(self):
        self.start_loading("Generating combined labels...")
        try:
            sections = self.collect_batch_sections()
        except Exception as exc:
            self.stop_loading()
            messagebox.showerror(APP_NAME, str(exc))
            return

        output_path = filedialog.asksaveasfilename(
            title="Save Combined Label Excel File",
            initialfile=self.first_batch_output_filename("Labels"),
            defaultextension=".xlsx",
            filetypes=[("Excel Workbook", "*.xlsx")],
        )

        if not output_path:
            self.stop_loading()
            return

        try:
            self.pulse_loading()
            write_combined_output(output_path, sections)
        except Exception as exc:
            self.stop_loading()
            messagebox.showerror(
                APP_NAME,
                f"Could not create combined file.\\n\\n{exc}",
            )
            return

        self.stop_loading()

        total = sum(len(section["labels"]) for section in sections)
        self.status_text.set(f"Created {total} combined cable label(s).")
        messagebox.showinfo(
            APP_NAME,
            f"Finished.\\n\\nCreated {total} combined cable label(s).",
        )

    def generate_batch_printable(self):
        self.start_loading("Generating combined printable labels...")
        try:
            sections = self.collect_batch_sections()
        except Exception as exc:
            self.stop_loading()
            messagebox.showerror(APP_NAME, str(exc))
            return

        output_path = filedialog.asksaveasfilename(
            title="Save Combined Printable Label File",
            initialfile=self.first_batch_output_filename("Printable Labels"),
            defaultextension=".xlsx",
            filetypes=[("Excel Workbook", "*.xlsx")],
        )

        if not output_path:
            self.stop_loading()
            return

        try:
            self.pulse_loading()
            write_combined_printable_output(output_path, sections)
        except Exception as exc:
            self.stop_loading()
            messagebox.showerror(
                APP_NAME,
                f"Could not create combined printable file.\\n\\n{exc}",
            )
            return

        self.stop_loading()

        total = sum(len(section["labels"]) for section in sections)
        self.status_text.set(f"Created {total} combined printable label(s).")
        messagebox.showinfo(
            APP_NAME,
            f"Finished.\\n\\nCreated {total} combined printable label(s).",
        )


    def generate_batch_easymark(self):
        self.start_loading("Generating combined Easy-Mark file...")
        try:
            sections = self.collect_batch_sections()
        except Exception as exc:
            self.stop_loading()
            messagebox.showerror(APP_NAME, str(exc))
            return

        output_path = filedialog.asksaveasfilename(
            title="Save Combined Easy-Mark File",
            initialfile=self.first_batch_output_filename("Easy-Mark"),
            defaultextension=".xlsx",
            filetypes=[("Excel Workbook", "*.xlsx")],
        )

        if not output_path:
            self.stop_loading()
            return

        try:
            self.pulse_loading()
            write_combined_easymark_output(output_path, sections)
        except Exception as exc:
            self.stop_loading()
            messagebox.showerror(
                APP_NAME,
                f"Could not create Easy-Mark file.\\n\\n{exc}",
            )
            return

        self.stop_loading()

        total = sum(len(section["labels"]) for section in sections)
        self.status_text.set(f"Created {total} combined Easy-Mark label(s).")
        messagebox.showinfo(
            APP_NAME,
            f"Finished.\\n\\nCreated {total} combined Easy-Mark label(s).",
        )

    def select_file(self):
        selected = filedialog.askopenfilename(
            title="Select Excel Cable File",
            filetypes=[
                ("Excel Workbooks", "*.xlsx *.xlsm"),
                ("Excel Workbook", "*.xlsx"),
                ("Excel Macro-Enabled Workbook", "*.xlsm"),
                ("All Files", "*.*"),
            ],
        )

        if not selected:
            return

        self.start_loading("Loading workbook...")
        self.file_path.set(selected)

        try:
            overview = self.get_cached_workbook_overview(selected)
            sheet_names = list(
                overview.get("sheet_names", [])
            )

            if not sheet_names:
                raise RuntimeError(
                    "The workbook does not contain any worksheets."
                )

            default_sheet = sheet_names[0]

            for candidate in sheet_names:
                info = overview.get("sheets", {}).get(
                    candidate,
                    {},
                )
                columns = info.get("auto_columns", {})

                if all(
                    columns.get(key)
                    for key in (
                        "a_position",
                        "z_position",
                        "cable_type",
                    )
                ):
                    default_sheet = candidate
                    break

            self.sheet_combo["values"] = sheet_names
            self.sheet_name.set(default_sheet)
            self.setup_column_mapping()

        except Exception as exc:
            self.sheet_combo["values"] = []
            self.sheet_name.set("")
            self.status_text.set(
                "Unable to read the selected workbook."
            )
            messagebox.showerror(APP_NAME, str(exc))
        finally:
            self.stop_loading()


    def on_sheet_selected(self, event=None):
        self.start_loading("Loading worksheet...")
        try:
            self.setup_column_mapping()
            self.pulse_loading()
        finally:
            self.stop_loading()

    def on_mapping_changed(self, event=None):
        self.start_loading("Updating mapping...")
        try:
            self.refresh_sheet_status()
            self.load_output_preview()
            self.pulse_loading()
        finally:
            self.stop_loading()

    def column_choice(self, col_num, header_value):
        letter = get_column_letter(col_num)
        header = clean_value(header_value)
        if not header:
            header = "(blank)"
        return f"{letter} - {header}"

    def choice_to_column(self, choice):
        choice = clean_value(choice)

        if not choice or choice == "(None)":
            return None

        letter = choice.split(" - ", 1)[0].strip().upper()

        col_num = 0
        for char in letter:
            if not ("A" <= char <= "Z"):
                return None
            col_num = (col_num * 26) + (ord(char) - ord("A") + 1)

        return col_num or None

    def get_selected_columns(self):
        columns = {
            "a_position": self.choice_to_column(self.a_column.get()),
            "z_position": self.choice_to_column(self.z_column.get()),
            "cable_type": self.choice_to_column(self.cable_type_column.get()),
        }

        path_col = self.choice_to_column(self.path_column.get())
        if path_col:
            columns["path"] = path_col

        if not all(
            columns.get(key)
            for key in ("a_position", "z_position", "cable_type")
        ):
            raise RuntimeError(
                "Please select columns for A Position, Z Position, and Cable Type."
            )

        return columns

    def setup_column_mapping(self):
        source = self.file_path.get().strip()
        sheet_name = self.sheet_name.get().strip()

        if not source or not sheet_name:
            return

        try:
            info = self.get_cached_sheet_info(
                source,
                sheet_name,
                include_preview=False,
            )

            header_row = info.get("header_row")
            auto_columns = info.get("auto_columns", {})

            if not auto_columns:
                self.current_header_row = None
                self.current_auto_columns = {}
                self.status_text.set(
                    f'{sheet_name}: required columns not detected.'
                )
                return

            self.current_header_row = header_row
            self.current_auto_columns = dict(auto_columns)

            choices = list(info.get("choices", []))
            choice_by_col = dict(
                info.get("choice_by_col", {})
            )

            for key, combo in self.mapping_combos.items():
                combo_values = list(choices)

                if key == "path":
                    combo_values = ["(None)"] + combo_values

                combo["values"] = combo_values

                detected_col = auto_columns.get(key)

                if detected_col:
                    combo.set(
                        choice_by_col.get(
                            detected_col,
                            "",
                        )
                    )
                elif key == "path":
                    combo.set("(None)")
                else:
                    combo.set("")

            self.refresh_sheet_status()
            self.refresh_previews()

        except Exception as exc:
            self.status_text.set(
                "Unable to read the selected worksheet."
            )
            messagebox.showerror(APP_NAME, str(exc))


    def refresh_sheet_status(self):
        source = self.file_path.get().strip()
        sheet_name = self.sheet_name.get().strip()

        if not source or not sheet_name or not self.current_header_row:
            return

        try:
            columns = self.get_selected_columns()

            labels = self.cached_extract_labels(
                source,
                sheet_name,
                header_row_override=self.current_header_row,
                columns_override=columns,
                progress_callback=self.pulse_loading,
            )

            detected = []
            for key, label in (
                ("a_position", "A Position"),
                ("cable_type", "Cable Type"),
                ("z_position", "Z Position"),
                ("path", "Path"),
            ):
                col_num = columns.get(key)
                if col_num:
                    detected.append(
                        f"{label}: {get_column_letter(col_num)}"
                    )
                else:
                    detected.append(f"{label}: not found")

            self.status_text.set(
                f'{sheet_name}: {len(labels)} cable(s) | '
                + " | ".join(detected)
            )

        except Exception as exc:
            self.status_text.set(str(exc))


    def generate_labels(self):
        source = self.file_path.get().strip()

        if not source:
            messagebox.showwarning(APP_NAME, "Please select an Excel file first.")
            return

        if not self.sheet_name.get().strip():
            messagebox.showwarning(APP_NAME, "Please select a worksheet.")
            return

        if not os.path.isfile(source):
            messagebox.showerror(APP_NAME, "The selected Excel file no longer exists.")
            return

        try:
            start_number = int(self.starting_cable_number.get().strip())
            if start_number < 0:
                raise ValueError
        except ValueError:
            messagebox.showwarning(
                APP_NAME,
                "Starting Cable Number must be a whole number of 0 or greater.",
            )
            return

        try:
            columns = self.get_selected_columns()
            labels = self.cached_extract_labels(
                source,
                self.sheet_name.get().strip(),
                header_row_override=self.current_header_row,
                columns_override=columns,
                progress_callback=self.pulse_loading,
            )
        except Exception as exc:
            messagebox.showerror(APP_NAME, str(exc))
            return

        if not labels:
            messagebox.showwarning(
                APP_NAME,
                "No cable labels were found in the selected workbook.",
            )
            return

        source_path = os.path.abspath(source)
        source_folder = os.path.dirname(source_path)
        source_name = os.path.splitext(os.path.basename(source_path))[0]
        worksheet_name = self.sanitize_filename_part(
            self.sheet_name.get().strip()
        )
        suggested_name = (
            f"{worksheet_name}_{source_name}_labels.xlsx"
            if worksheet_name
            else f"{source_name}_labels.xlsx"
        )

        output_path = filedialog.asksaveasfilename(
            title="Save Label Excel File",
            initialdir=source_folder,
            initialfile=suggested_name,
            defaultextension=".xlsx",
            filetypes=[("Excel Workbook", "*.xlsx")],
        )

        if not output_path:
            return

        self.start_loading("Generating labels...")

        try:
            self.pulse_loading()
            write_output(
                output_path=output_path,
                labels=labels,
                project_number=self.project_number.get().strip(),
                starting_cable_number=start_number,
                identifier=self.identifier.get().strip(),
                other=self.other.get().strip(),
            )
        except PermissionError:
            self.stop_loading()
            messagebox.showerror(
                APP_NAME,
                "The output file could not be saved.\n\n"
                "If it is already open in Excel, close it and try again.",
            )
            return
        except Exception as exc:
            messagebox.showerror(
                APP_NAME,
                f"Could not create the label file.\n\n{exc}",
            )
            return

        self.stop_loading()

        self.status_text.set(
            f"Created {len(labels)} label(s): {os.path.basename(output_path)}"
        )

        messagebox.showinfo(
            APP_NAME,
            f"Finished.\n\nCreated {len(labels)} cable label(s).",
        )

    def generate_printable_labels(self):
        source = self.file_path.get().strip()

        if not source:
            messagebox.showwarning(APP_NAME, "Please select an Excel file first.")
            return

        if not self.sheet_name.get().strip():
            messagebox.showwarning(APP_NAME, "Please select a worksheet.")
            return

        if not os.path.isfile(source):
            messagebox.showerror(APP_NAME, "The selected Excel file no longer exists.")
            return

        try:
            start_number = int(self.starting_cable_number.get().strip())
            if start_number < 0:
                raise ValueError
        except ValueError:
            messagebox.showwarning(
                APP_NAME,
                "Starting Cable Number must be a whole number of 0 or greater.",
            )
            return

        try:
            columns = self.get_selected_columns()
            labels = self.cached_extract_labels(
                source,
                self.sheet_name.get().strip(),
                header_row_override=self.current_header_row,
                columns_override=columns,
                progress_callback=self.pulse_loading,
            )
        except Exception as exc:
            messagebox.showerror(APP_NAME, str(exc))
            return

        if not labels:
            messagebox.showwarning(
                APP_NAME,
                "No cable labels were found in the selected workbook.",
            )
            return

        source_path = os.path.abspath(source)
        source_folder = os.path.dirname(source_path)
        source_name = os.path.splitext(os.path.basename(source_path))[0]
        worksheet_name = self.sanitize_filename_part(
            self.sheet_name.get().strip()
        )
        suggested_name = (
            f"{worksheet_name}_{source_name}_printable_labels.xlsx"
            if worksheet_name
            else f"{source_name}_printable_labels.xlsx"
        )

        output_path = filedialog.asksaveasfilename(
            title="Save Printable Label Excel File",
            initialdir=source_folder,
            initialfile=suggested_name,
            defaultextension=".xlsx",
            filetypes=[("Excel Workbook", "*.xlsx")],
        )

        if not output_path:
            return

        self.start_loading("Generating printable labels...")

        try:
            self.pulse_loading()
            write_printable_output(
                output_path=output_path,
                labels=labels,
                project_number=self.project_number.get().strip(),
                starting_cable_number=start_number,
                identifier=self.identifier.get().strip(),
                other=self.other.get().strip(),
            )
        except PermissionError:
            self.stop_loading()
            messagebox.showerror(
                APP_NAME,
                "The output file could not be saved.\n\n"
                "If it is already open in Excel, close it and try again.",
            )
            return
        except Exception as exc:
            messagebox.showerror(
                APP_NAME,
                f"Could not create the printable label file.\n\n{exc}",
            )
            return

        self.stop_loading()

        self.status_text.set(
            f"Created {len(labels)} printable label(s): "
            f"{os.path.basename(output_path)}"
        )

        messagebox.showinfo(
            APP_NAME,
            f"Finished.\n\nCreated {len(labels)} printable cable label(s).",
        )

    def generate_easymark_file(self):
        source = self.file_path.get().strip()

        if not source:
            messagebox.showwarning(APP_NAME, "Please select an Excel file first.")
            return

        if not self.sheet_name.get().strip():
            messagebox.showwarning(APP_NAME, "Please select a worksheet.")
            return

        if not os.path.isfile(source):
            messagebox.showerror(APP_NAME, "The selected Excel file no longer exists.")
            return

        try:
            start_number = int(self.starting_cable_number.get().strip())
            if start_number < 0:
                raise ValueError
        except ValueError:
            messagebox.showwarning(
                APP_NAME,
                "Starting Cable Number must be a whole number of 0 or greater.",
            )
            return

        try:
            columns = self.get_selected_columns()
            labels = self.cached_extract_labels(
                source,
                self.sheet_name.get().strip(),
                header_row_override=self.current_header_row,
                columns_override=columns,
                progress_callback=self.pulse_loading,
            )
        except Exception as exc:
            messagebox.showerror(APP_NAME, str(exc))
            return

        if not labels:
            messagebox.showwarning(
                APP_NAME,
                "No cable labels were found in the selected workbook.",
            )
            return

        source_path = os.path.abspath(source)
        source_folder = os.path.dirname(source_path)
        source_name = os.path.splitext(os.path.basename(source_path))[0]
        worksheet_name = self.sanitize_filename_part(
            self.sheet_name.get().strip()
        )
        suggested_name = (
            f"{worksheet_name}_{source_name}_easymark.xlsx"
            if worksheet_name
            else f"{source_name}_easymark.xlsx"
        )

        output_path = filedialog.asksaveasfilename(
            title="Save Easy-Mark Excel File",
            initialdir=source_folder,
            initialfile=suggested_name,
            defaultextension=".xlsx",
            filetypes=[("Excel Workbook", "*.xlsx")],
        )

        if not output_path:
            return

        self.start_loading("Generating Easy-Mark file...")

        try:
            self.pulse_loading()
            write_easymark_output(
                output_path=output_path,
                labels=labels,
                project_number=self.project_number.get().strip(),
                starting_cable_number=start_number,
                identifier=self.identifier.get().strip(),
                other=self.other.get().strip(),
            )
        except PermissionError:
            self.stop_loading()
            messagebox.showerror(
                APP_NAME,
                "The output file could not be saved.\n\n"
                "If it is already open in Excel, close it and try again.",
            )
            return
        except Exception as exc:
            messagebox.showerror(
                APP_NAME,
                f"Could not create the Easy-Mark file.\n\n{exc}",
            )
            return

        self.stop_loading()

        self.status_text.set(
            f"Created {len(labels)} Easy-Mark label(s): "
            f"{os.path.basename(output_path)}"
        )

        messagebox.showinfo(
            APP_NAME,
            f"Finished.\n\nCreated {len(labels)} Easy-Mark label(s).",
        )

    def clear_fields(self):
        self.file_path.set("")
        self.sheet_name.set("")
        self.sheet_combo["values"] = []

        self.a_column.set("")
        self.z_column.set("")
        self.cable_type_column.set("")
        self.path_column.set("")
        self.current_header_row = None
        self.current_auto_columns = {}

        for combo in self.mapping_combos.values():
            combo["values"] = []

        self.project_number.set("")
        self.starting_cable_number.set("1")
        self.identifier.set("")
        self.other.set("")
        self.clear_tree(self.output_preview)
        self.status_text.set("Select an Excel file to begin.")


def main():
    set_windows_app_identity()

    # className sets the X11/Wayland WM_CLASS used by Linux desktop
    # environments for taskbar grouping/name matching.
    root = tk.Tk(className="MTGCableTool")
    root.title(APP_NAME)
    root.iconname(APP_NAME)

    try:
        style = ttk.Style()
        if "clam" in style.theme_names():
            style.theme_use("clam")
    except Exception:
        pass

    LabelGeneratorApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
